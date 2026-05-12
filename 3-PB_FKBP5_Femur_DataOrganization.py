import os
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime

# --- CONFIGURATION ---
base_dir = r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending"
raw_data_root = os.path.join(
    base_dir, "Force-Displacement Raw Files", "FKBP5Null_Tibia_11226")
master_file = r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending\FKBP5_3-PointBendingFemurMaster.xlsx"
measurement_file = r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending\Measurement Files\FKBP5Null_Tibia+Femur_11226.xlsx"
output_file = r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending\FKBP5_Femur_Compiled.csv"


def sync_femur_to_master(all_analysis_files, master_file, measurement_file):
    df_meas_all = pd.read_excel(measurement_file, sheet_name=None)
    wb = openpyxl.load_workbook(master_file)
    folder_to_file_map = {}

    def get_date(path_obj):
        try:
            return datetime.strptime(path_obj.stem.split('_')[-1], "%m%d%y")
        except:
            return datetime.min

    for file_path in all_analysis_files:
        folder_name = file_path.parent.name
        if folder_name not in folder_to_file_map:
            folder_to_file_map[folder_name] = file_path
        else:
            if get_date(file_path) > get_date(folder_to_file_map[folder_name]):
                folder_to_file_map[folder_name] = file_path

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ["summary", "notes", "calculations"]:
            continue
        if sheet_name not in folder_to_file_map:
            print(f"Skipping {sheet_name}: No matching folder found.")
            continue

        print(
            f"Syncing Femur Sheet '{sheet_name}' using: {folder_to_file_map[sheet_name].name}")
        ws = wb[sheet_name]
        df_mach = pd.read_excel(folder_to_file_map[sheet_name])
        df_mach['Clean_Filename'] = df_mach['Filename'].str.replace(
            '.txt', '', regex=False)
        df_meas = df_meas_all.get(sheet_name, pd.DataFrame())

        for row in range(2, ws.max_row + 1):
            for start_col in [1, 10]:  # Column A and Column J
                mouse_id = ws.cell(row=row, column=start_col).value
                if not mouse_id:
                    continue

                femur_search_key = f"{str(mouse_id).strip()}_Femur"

                if not df_meas.empty:
                    m_row = df_meas[df_meas.iloc[:, 0].astype(
                        str).str.strip() == femur_search_key]
                    if not m_row.empty:
                        ws.cell(row=row, column=start_col +
                                1).value = m_row.iloc[0, 1]
                        ws.cell(row=row, column=start_col +
                                2).value = m_row.iloc[0, 8]
                        ws.cell(row=row, column=start_col +
                                3).value = m_row.iloc[0, 12]

                mach_row = df_mach[df_mach['Clean_Filename']
                                   == femur_search_key]
                if not mach_row.empty:
                    ws.cell(row=row, column=start_col +
                            4).value = mach_row.iloc[0]['Max_Load_N']
                    ws.cell(row=row, column=start_col +
                            5).value = mach_row.iloc[0]['Stiffness_N_per_mm']
                    ws.cell(row=row, column=start_col +
                            6).value = mach_row.iloc[0]['Energy_to_Failure_Nmm']
                    ws.cell(row=row, column=start_col +
                            7).value = mach_row.iloc[0]['Displacement_at_Failure_mm']

    wb.save(master_file)
    print("Femur Master update complete.")


def parse_mouse_code(code):
    try:
        base_code = str(code).split('_')[0]
        parts = base_code.split('.')
        genotype = parts[0]
        age = parts[1]
        sex_id = parts[2]
        sex = "Male" if sex_id.startswith('M') else "Female"
        mouse_num = sex_id[1:]
        return genotype, age, sex, mouse_num
    except:
        return None, None, None, None


def generate_prism_tables(df):
    """Generates the grouped CSV files for GraphPad Prism."""
    metrics = [
        'Avg. Femur Length', 'Avg. Femur Diameter', 'Avg. Femur Thickness',
        'Maximum Load', 'Stiffness', 'Energy to Failure', 'Displacement at Failure'
    ]
    unique_ages = df['Age_Extracted'].unique()
    sort_priority = ['Wildtype', 'Mutant', 'Heterozygous']

    # folder_genotype = os.path.join(base_dir, "Femur_Analysis_By_Genotype")
    folder_lineage = os.path.join(base_dir, "Femur_Analysis_By_Lineage")
    # os.makedirs(folder_genotype, exist_ok=True)
    os.makedirs(folder_lineage, exist_ok=True)

    for metric in metrics:
        for age in unique_ages:
            if pd.isna(age):
                continue
            for sex in ['Male', 'Female']:
                base_subset = df[(df['Sex_Extracted'] == sex) & (
                    df['Age_Extracted'] == age)].copy()
                if base_subset.empty:
                    continue

                clean_metric = metric.replace(" ", "_").replace(".", "")

                # # SET 1: By Genotype
                # gen_subset = base_subset[base_subset['Progeny_Group'].isin(
                #     sort_priority)].copy()
                # if not gen_subset.empty:
                #     table_gen = gen_subset.pivot_table(
                #         index='ID_Num', columns='Progeny_Group', values=metric)
                #     existing_cols = [
                #         c for c in sort_priority if c in table_gen.columns]
                #     table_gen = table_gen[existing_cols]
                #     fname = f"{sex}_{age}Wks_{clean_metric}.csv"
                #     table_gen.to_csv(os.path.join(folder_genotype, fname))

                # SET 2: By Lineage
                table_lin = base_subset.pivot_table(
                    index='ID_Num', columns='Progeny_Group', values=metric)

                def lin_sort(c):
                    for i, g in enumerate(sort_priority):
                        if c.startswith(g):
                            return (i, c)
                    return (99, c)
                sorted_cols = sorted(table_lin.columns, key=lin_sort)
                table_lin = table_lin[sorted_cols]
                table_lin.to_csv(os.path.join(
                    folder_lineage, f"{sex}_{age}Wks_{clean_metric}.csv"))


def process_all_sheets(path):
    excel_data = pd.ExcelFile(path)
    all_compiled_data = []

    for sheet in excel_data.sheet_names:
        if sheet.lower() in ["summary", "notes", "calculations"]:
            continue

        raw_df = pd.read_excel(path, sheet_name=sheet, header=None, skiprows=1)
        headers = ['Mouse Code', 'Avg. Femur Length', 'Avg. Femur Diameter', 'Avg. Femur Thickness',
                   'Maximum Load', 'Stiffness', 'Energy to Failure', 'Displacement at Failure']

        males = raw_df.iloc[:, 0:8].copy()
        females = raw_df.iloc[:, 9:17].copy()
        males.columns = headers
        females.columns = headers

        combined = pd.concat([males, females], ignore_index=True).dropna(
            subset=['Mouse Code'])
        combined['Progeny_Group'] = sheet
        all_compiled_data.append(combined)

    final_df = pd.concat(all_compiled_data, ignore_index=True)
    final_df[['Genotype', 'Age_Extracted', 'Sex_Extracted', 'ID_Num']] = \
        final_df['Mouse Code'].apply(lambda x: pd.Series(parse_mouse_code(x)))

    # Generate the Prism CSVs
    generate_prism_tables(final_df)

    final_df.to_csv(output_file, index=False)
    print(f"Compiled Master CSV saved to: {output_file}")


if __name__ == "__main__":
    try:
        root_path = Path(raw_data_root)
        all_files = list(root_path.rglob("Fz_Displacement_Analysis_*.xlsx"))
        sync_femur_to_master(all_files, master_file, measurement_file)
        process_all_sheets(master_file)
        print("\nFemur Workflow Complete!")
    except Exception as e:
        print(f"\nAn error occurred: {e}")
