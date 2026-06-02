import os
import glob
import re
from datetime import datetime
from pathlib import Path

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

matplotlib.use("Agg")

from bending_core import (
    read_bending_txt, dominant_linear_region,
    TOE_LOAD_FRACTION, LINEAR_WINDOW_POINTS, MIN_R2,
    DISPLACEMENT_LIMIT, NEAR_ZERO_THRESHOLD, DROP_THRESHOLD_FRACTION, PEAK_CANDIDATE_FRACTION,
)

# ===========================================================
# DEFAULT CONFIGURATION & CONSTANTS
# ===========================================================
FILE_GLOB_PATTERN = "*.txt"
SAVE_PNG_DPI = 100


def parse_id_and_group(text_string, group_map):
    if not text_string or pd.isna(text_string):
        return None, None, None

    text_string = str(text_string).strip()
    prefixes = list(group_map.keys())

    # Improved regex: allows flexible separators between the prefix and the number
    pattern = rf"\b({'|'.join(prefixes)})\s*[-_.]?\s*(\d+)"
    match = re.search(pattern, text_string, re.IGNORECASE)

    if match:
        pref = match.group(1).upper()
        num_id = match.group(2)
        val = group_map.get(pref, pref)
        group_name = val.get('group', pref) if isinstance(val, dict) else val
        return pref, num_id, group_name

    return None, None, None


def _build_standardized_id(pref, num_id, source_text):
    """Returns ID like 'Z1M' or 'EP1' — preserves trailing M/F sex letter when present."""
    m = re.search(r'\d+([MF])(?:[^a-zA-Z]|$)', str(source_text), re.IGNORECASE)
    suffix = m.group(1).upper() if m else ""
    return f"{pref}{num_id}{suffix}"


def _is_mixed_group_map(group_map):
    return any(isinstance(v, dict) for v in group_map.values())


def _get_prefix_sex(pref, group_map):
    val = group_map.get(pref)
    return val.get('sex') if isinstance(val, dict) else None


def _resolve_sex(pref, group_map, mouse_code):
    """Returns the sex for a Mixed-mode animal, deducing from the code when requested."""
    stored = _get_prefix_sex(pref, group_map)
    if stored == _DEDUCE_SEX:
        return _extract_sex_from_code(mouse_code)
    return stored


_KNOWN_BONES = ['Femur', 'Tibia', 'Humerus']
_DEDUCE_SEX = 'Deduce from code'


def _extract_sex_from_code(text_string):
    """Returns 'Male', 'Female', or None by reading trailing M/F before a non-letter."""
    m = re.search(r'\d+([MF])(?:[^a-zA-Z]|$)', str(text_string), re.IGNORECASE)
    if m:
        return 'Male' if m.group(1).upper() == 'M' else 'Female'
    return None


def _detect_bones_from_measurement(measurement_path, sheet_name):
    """Returns list of bone types found via _BoneName in mouse code column, or [] if none."""
    if not measurement_path or not os.path.exists(measurement_path):
        return []
    try:
        df = pd.read_excel(measurement_path, sheet_name=sheet_name if sheet_name else 0, header=0)
        return [b for b in _KNOWN_BONES
                if df.iloc[:, 0].astype(str).str.contains(f'_{b}', case=False, na=False).any()]
    except Exception:
        return []


# ===========================================================
# PART 1: BENDING DATA ANALYZER FUNCTIONS
# ===========================================================

def run_batch_bending_analysis(input_folder, group_map):
    print(f"\n--- Processing Raw Data in Folder: {input_folder} ---")
    date_str = datetime.now().strftime("%m%d%y")

    plot_folder = os.path.join(input_folder, "Fz_Displacement_Analysis")
    os.makedirs(plot_folder, exist_ok=True)

    excel_filename = f"Fz_Displacement_Analysis_{date_str}.xlsx"
    txt_files = glob.glob(os.path.join(input_folder, FILE_GLOB_PATTERN))
    if not txt_files:
        return None

    results = []

    for path in txt_files:
        base_name = os.path.basename(path).replace(".txt", "")

        pref, num_id, g_name = parse_id_and_group(base_name, group_map)
        if not pref:
            print(
                f"Skipping file {base_name}: Doesn't match active group configuration prefixes.")
            continue

        try:
            df = read_bending_txt(path)
            displacement = df["Position (z), mm"].values
            load = df["Fz, N"].values

            load_smooth = np.convolve(load, np.ones(3) / 3, mode="same")
            valid_range_mask = displacement <= DISPLACEMENT_LIMIT

            load_for_peak = load_smooth[valid_range_mask]
            constrained_max = np.max(load_for_peak) if len(
                load_for_peak) > 0 else np.max(load_smooth)

            candidates = np.where(
                (load_smooth >= PEAK_CANDIDATE_FRACTION * constrained_max) & (valid_range_mask))[0]
            max_idx = candidates[np.argmax(load[candidates])] if len(
                candidates) > 0 else np.argmax(load_smooth)

            max_load = load[max_idx]
            post_max_load = load[max_idx:]
            drop_threshold_value = max_load * DROP_THRESHOLD_FRACTION

            zero_drop_indices = np.where(
                post_max_load <= NEAR_ZERO_THRESHOLD)[0]
            if len(zero_drop_indices) > 0:
                fail_idx = max_idx + zero_drop_indices[0]
            else:
                has_dropped_indices = np.where(
                    post_max_load < drop_threshold_value)[0]
                if len(has_dropped_indices) > 0:
                    search_start = has_dropped_indices[0]
                    search_load = post_max_load[search_start:]
                    diff_post = np.diff(np.convolve(
                        search_load, np.ones(3) / 3, mode="same"))
                    local_min_indices = np.where(diff_post >= 0)[0]
                    fail_idx = (
                        max_idx
                        + search_start
                        + (local_min_indices[0] if len(local_min_indices) > 0 else len(search_load) - 1)
                    )
                else:
                    fail_idx = len(load) - 1

            pre_max_load = load[:max_idx]
            toe_mask = pre_max_load >= (TOE_LOAD_FRACTION * max_load)
            toe_indices = np.where(toe_mask)[0]

            start_idx = toe_indices[0] if len(toe_indices) > 0 else 0
            adj_disp = displacement - displacement[start_idx]

            disp_at_failure = adj_disp[fail_idx]
            load_at_failure = load[fail_idx]

            disp_slope_candidates = displacement[:max_idx][toe_mask] - \
                displacement[start_idx]
            load_slope_candidates = pre_max_load[toe_mask]

            if len(disp_slope_candidates) < LINEAR_WINDOW_POINTS:
                stiffness, intercept, idx0, idx1 = np.nan, np.nan, 0, 0
            else:
                stiffness, intercept, idx0, idx1 = dominant_linear_region(
                    disp_slope_candidates, load_slope_candidates, window=LINEAR_WINDOW_POINTS
                )

            if hasattr(np, "trapezoid"):
                energy = np.trapezoid(
                    load[start_idx: fail_idx + 1], adj_disp[start_idx: fail_idx + 1])
            else:
                energy = np.trapz(
                    load[start_idx: fail_idx + 1], adj_disp[start_idx: fail_idx + 1])

            plt.figure(figsize=(7, 5))
            plt.plot(adj_disp, load, color="black", label="Data")
            if (idx1 - idx0) > 5 and np.isfinite(stiffness):
                x_lin = disp_slope_candidates[idx0:idx1]
                plt.plot(x_lin, stiffness * x_lin + intercept,
                         color="red", label=f"Stiffness: {stiffness:.2f} N/mm")

            plt.scatter(adj_disp[max_idx], max_load,
                        color="green", label="Max Load")
            plt.scatter(disp_at_failure, load_at_failure,
                        color="purple", label="Failure")
            plt.axvline(x=0, color="blue", linestyle="--", label="Toe End")
            plt.fill_between(adj_disp[start_idx: fail_idx + 1],
                             load[start_idx: fail_idx + 1], alpha=0.2, color="orange")
            plt.title(base_name)
            plt.legend()
            plt.savefig(os.path.join(
                plot_folder, f"{base_name}.png"), dpi=SAVE_PNG_DPI)
            plt.close()

            bone_in_name = next(
                (b for b in _KNOWN_BONES if b.lower() in base_name.lower()), None)
            results.append(
                {
                    "Standardized_ID": _build_standardized_id(pref, num_id, base_name),
                    "Bone": bone_in_name,
                    "Max_Load_N": round(max_load, 4),
                    "Stiffness_N_per_mm": round(stiffness, 4),
                    "Energy_to_Failure_Nmm": round(energy, 4),
                    "Displacement_at_Failure_mm": round(disp_at_failure, 4),
                }
            )
        except Exception as e:
            print(f"Error processing mechanics for {base_name}: {e}")

    if not results:
        return None

    output_excel_path = os.path.join(input_folder, excel_filename)
    pd.DataFrame(results).to_excel(output_excel_path, index=False)
    return output_excel_path


# ===========================================================
# PART 2: MASTER MERGING & CONSOLIDATION
# ===========================================================

def sync_data_to_master(analysis_excel_path, master_file, measurement_file, group_map, bone_filter=None):
    if not analysis_excel_path or not os.path.exists(analysis_excel_path):
        return

    df_mach = pd.read_excel(analysis_excel_path)
    df_meas = pd.DataFrame()

    if os.path.exists(measurement_file):
        try:
            df_meas = pd.read_excel(measurement_file)
        except Exception as e:
            print(f"Could not read measurement file: {e}")

    # Filter measurement rows to matching bone
    if bone_filter and not df_meas.empty:
        mask = df_meas.iloc[:, 0].astype(str).str.contains(f'_{bone_filter}', case=False, na=False)
        if mask.any():
            df_meas = df_meas[mask].reset_index(drop=True)

    # Filter mechanical results to matching bone
    if bone_filter and "Bone" in df_mach.columns:
        mach_bone = df_mach[df_mach["Bone"].fillna("").str.lower() == bone_filter.lower()]
        if not mach_bone.empty:
            df_mach = mach_bone

    def _find_meas_col(df, *names):
        for name in names:
            if name in df.columns:
                return name
        return None

    _meas_len_col   = _find_meas_col(df_meas, "Length")
    _meas_diam_col  = _find_meas_col(df_meas, "D_avg")
    _meas_thick_col = _find_meas_col(df_meas, "h_avg")

    wb = openpyxl.load_workbook(master_file, keep_links=True)
    ws = wb.active

    # === NEW: AUTO-POPULATE EMPTY MASTER LIST ===
    if ws.max_row <= 1:
        all_ids = set()

        # Grab IDs from mechanical logs
        if not df_mach.empty and "Standardized_ID" in df_mach.columns:
            for val in df_mach["Standardized_ID"].dropna():
                all_ids.add(str(val))

        # Grab IDs from the measurement file
        if not df_meas.empty:
            for idx, m_row in df_meas.iterrows():
                pref, num_id, _ = parse_id_and_group(m_row.iloc[0], group_map)
                if pref:
                    all_ids.add(_build_standardized_id(pref, num_id, str(m_row.iloc[0])))

        # Write sorted IDs into Column 1
        sorted_ids = sorted(list(all_ids))
        for i, mouse_id in enumerate(sorted_ids, start=2):
            ws.cell(row=i, column=1).value = mouse_id

    # Recalculate max_row now that it is populated
    max_r = ws.max_row
    # ============================================

    headers = [str(ws.cell(row=1, column=c).value).strip()
               for c in range(1, ws.max_column + 1)]

    def find_col(name_options):
        for opt in name_options:
            if opt in headers:
                return headers.index(opt) + 1
        return None

    col_len = find_col(["Length", "Avg. Length", "Avg_Length"]) or 2
    col_diam = find_col(["Diameter", "Avg. Diameter", "Avg_Diameter"]) or 3
    col_thick = find_col(["Thickness", "Avg. Thickness", "Avg_Thickness"]) or 4
    col_max = find_col(["Maximum Load", "Max_Load", "Max_Load_N"]) or 5
    col_stiff = find_col(["Stiffness", "Stiffness_N_per_mm"]) or 6
    col_energy = find_col(
        ["Energy to Failure", "Energy_to_Failure", "Energy_to_Failure_Nmm"]) or 7
    col_disp = find_col(["Displacement at Failure",
                        "Displacement_at_Failure", "Displacement_at_Failure_mm"]) or 8

    for row in range(2, max_r + 1):
        cell_val1 = ws.cell(row=row, column=1).value
        cell_val2 = ws.cell(row=row, column=2).value

        src_val = cell_val1
        pref, num_id, _ = parse_id_and_group(cell_val1, group_map)
        if not pref:
            src_val = cell_val2
            pref, num_id, _ = parse_id_and_group(cell_val2, group_map)

        if not pref:
            continue

        lookup_id = _build_standardized_id(pref, num_id, str(src_val or ''))

        if not df_meas.empty:
            for idx, m_row in df_meas.iterrows():
                m_pref, m_num, _ = parse_id_and_group(m_row.iloc[0], group_map)
                if m_pref and _build_standardized_id(m_pref, m_num, str(m_row.iloc[0])) == lookup_id:
                    ws.cell(row=row, column=col_len).value = (
                        m_row[_meas_len_col] if _meas_len_col else None)
                    ws.cell(row=row, column=col_diam).value = (
                        m_row[_meas_diam_col] if _meas_diam_col else None)
                    ws.cell(row=row, column=col_thick).value = (
                        m_row[_meas_thick_col] if _meas_thick_col else None)
                    break

        mach_row = df_mach[df_mach["Standardized_ID"] == lookup_id]
        if not mach_row.empty:
            ws.cell(
                row=row, column=col_max).value = mach_row.iloc[0]["Max_Load_N"]
            ws.cell(
                row=row, column=col_stiff).value = mach_row.iloc[0]["Stiffness_N_per_mm"]
            ws.cell(
                row=row, column=col_energy).value = mach_row.iloc[0]["Energy_to_Failure_Nmm"]
            ws.cell(
                row=row, column=col_disp).value = mach_row.iloc[0]["Displacement_at_Failure_mm"]

    wb.save(master_file)

# ===========================================================
# PART 3: SEGREGATED SUB-STUDY CSV GENERATION
# ===========================================================


def generate_segregated_csvs(master_path, export_dir, group_map):
    if master_path.endswith(".csv"):
        df = pd.read_csv(master_path)
    else:
        df = pd.read_excel(master_path)

    headers = [str(c).strip() for c in df.columns]

    def find_col_name(name_options):
        for opt in name_options:
            if opt in headers:
                return opt
            for h in headers:
                if opt.lower() in h.lower():
                    return h
        return None

    metric_mapping = {
        "Avg_Length": find_col_name(["Length", "Avg_Length", "Avg. Length"]),
        "Avg_Diameter": find_col_name(["Diameter", "Avg_Diameter", "Avg. Diameter"]),
        "Avg_Thickness": find_col_name(["Thickness", "Avg_Thickness", "Avg. Thickness"]),
        "Max_Load": find_col_name(["Max_Load", "Maximum Load", "Max Load", "Max_Load_N"]),
        "Stiffness": find_col_name(["Stiffness", "Stiffness_N_per_mm", "Stiffness (N/mm)"]),
        "Energy_to_Failure": find_col_name(["Energy", "Energy_to_Failure", "Energy to Failure"]),
        "Displacement_at_Failure": find_col_name(["Displacement", "Displacement_at_Failure"]),
    }

    is_mixed = _is_mixed_group_map(group_map)
    parsed_rows = []

    for idx, row in df.iterrows():
        matched_pref, matched_num, matched_gname, matched_src = None, None, None, None

        for col_idx in range(min(3, len(df.columns))):
            p, n, g = parse_id_and_group(row.iloc[col_idx], group_map)
            if p:
                matched_pref, matched_num, matched_gname = p, n, g
                matched_src = str(row.iloc[col_idx])
                break

        if not matched_pref:
            continue

        row_data = {
            "ID_Num": _build_standardized_id(matched_pref, matched_num, matched_src),
            "Group": matched_gname,
            "Sex": _resolve_sex(matched_pref, group_map, matched_src) if is_mixed else None,
        }

        for metric_key, actual_col in metric_mapping.items():
            row_data[metric_key] = row[actual_col] if actual_col else np.nan

        parsed_rows.append(row_data)

    if not parsed_rows:
        print("Notice: No matching cohort IDs could be extracted from master rows.")
        return

    df_parsed = pd.DataFrame(parsed_rows)
    os.makedirs(export_dir, exist_ok=True)

    def write_pivot(df_sub, col_key, metric_name, out_dir, expected_cols):
        if df_sub[metric_name].isna().all():
            return
        pivot_df = df_sub.pivot_table(
            index="ID_Num", columns=col_key, values=metric_name, aggfunc="first")
        pivot_df = pivot_df.reindex(columns=expected_cols)
        pivot_df = pivot_df.sort_index()
        out_path = os.path.join(out_dir, f"{metric_name}.csv")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write("sep=,\n")
        pivot_df.to_csv(out_path, mode="a", sep=",", index=True)
        print(f"Generated side-by-side table: {out_path}")

    if is_mixed:
        df_parsed["Group_Sex"] = df_parsed.apply(
            lambda r: f"{r['Group']}_{r['Sex']}" if r["Sex"] else r["Group"], axis=1)

        has_deduced = any(
            isinstance(v, dict) and v.get('sex') == _DEDUCE_SEX for v in group_map.values())

        if has_deduced:
            combined_cols = list(dict.fromkeys(df_parsed["Group_Sex"].dropna()))
        else:
            combined_cols = list(dict.fromkeys(
                f"{v.get('group', '')}_{v.get('sex', '')}" if isinstance(v, dict) else v
                for v in group_map.values()
            ))

        for metric_name in metric_mapping:
            write_pivot(df_parsed, "Group_Sex", metric_name, export_dir, combined_cols)

        for sex_label in ["Male", "Female"]:
            sex_dir = os.path.join(export_dir, sex_label)
            os.makedirs(sex_dir, exist_ok=True)
            sex_df = df_parsed[df_parsed["Sex"] == sex_label].copy()
            if sex_df.empty:
                continue
            if has_deduced:
                sex_groups = list(dict.fromkeys(sex_df["Group"].dropna()))
            else:
                sex_groups = list(dict.fromkeys(
                    v.get("group", "") for v in group_map.values()
                    if isinstance(v, dict) and v.get("sex") == sex_label
                ))
            for metric_name in metric_mapping:
                write_pivot(sex_df, "Group", metric_name, sex_dir, sex_groups)
    else:
        expected_columns = list(group_map.values())
        for metric_name in metric_mapping:
            write_pivot(df_parsed, "Group", metric_name, export_dir, expected_columns)


# ===========================================================
# PART 4: ANATOMICAL DIAMETER PARSING
# ===========================================================

def parse_anatomical_diameters_single_study(measurement_path, sheet_name, csv_out_dir, bone, sex, age, group_map, bone_filter=None):
    if not measurement_path or not os.path.exists(measurement_path):
        print("Anatomical diameters: measurement file not found, skipping.")
        return

    try:
        df = pd.read_excel(measurement_path, sheet_name=sheet_name if sheet_name else 0, header=0)
    except Exception as e:
        print(f"Anatomical diameters: could not read sheet '{sheet_name}': {e}")
        return

    if bone_filter:
        mask = df.iloc[:, 0].astype(str).str.contains(f'_{bone_filter}', case=False, na=False)
        if mask.any():
            df = df[mask].reset_index(drop=True)

    if bone == "Tibia":
        label_top, label_bottom = "Proximal_Tibia", "Distal_Tibia"
    else:
        label_top, label_bottom = f"Anteroposterior_{bone}", f"Mediolateral_{bone}"

    is_mixed = _is_mixed_group_map(group_map)

    cols_g1 = [c for c in ('D1', 'D2', 'D3') if c in df.columns]
    cols_g2 = [c for c in ('D4', 'D5', 'D6') if c in df.columns]

    if not cols_g1:
        print("Anatomical diameters: no D1/D2/D3 columns found in measurement file, skipping.")
        return
    if not cols_g2:
        print("Anatomical diameters: only 3 diameter measurements found — AP/ML split requires D1–D6. Skipping.")
        return

    rows = []
    for _, row in df.iterrows():
        pref, num_id, group_name = parse_id_and_group(row.iloc[0], group_map)
        if not pref:
            continue
        try:
            avg1 = pd.to_numeric(row[cols_g1], errors='coerce').mean()
            avg2 = pd.to_numeric(row[cols_g2], errors='coerce').mean()
        except Exception:
            continue
        if pd.isna(avg1) or pd.isna(avg2):
            continue
        rows.append({
            'ID_Num': _build_standardized_id(pref, num_id, str(row.iloc[0])),
            'Group': group_name,
            'Sex': _resolve_sex(pref, group_map, row.iloc[0]) if is_mixed else None,
            'Top_Val': max(avg1, avg2),
            'Bottom_Val': min(avg1, avg2),
        })

    if not rows:
        print("Anatomical diameters: no matching rows found in measurement file.")
        return

    master_df = pd.DataFrame(rows)
    folder_top = os.path.join(csv_out_dir, label_top)
    folder_bottom = os.path.join(csv_out_dir, label_bottom)
    os.makedirs(folder_top, exist_ok=True)
    os.makedirs(folder_bottom, exist_ok=True)

    age_clean = age.replace(" ", "_")

    def write_anat_pivot(df_sub, col_key, data_key, target_folder, label, file_prefix, expected_cols):
        pivot = df_sub.pivot_table(
            index='ID_Num', columns=col_key, values=data_key, aggfunc='first')
        pivot = pivot.reindex(columns=expected_cols)
        pivot = pivot.sort_index()
        out_path = os.path.join(target_folder, f"{file_prefix}_{label}.csv")
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write("sep=,\n")
        pivot.to_csv(out_path, mode='a', sep=',', index=True)
        print(f"Generated anatomical diameter table: {out_path}")

    if is_mixed:
        master_df["Group_Sex"] = master_df.apply(
            lambda r: f"{r['Group']}_{r['Sex']}" if r["Sex"] else r["Group"], axis=1)

        has_deduced = any(
            isinstance(v, dict) and v.get('sex') == _DEDUCE_SEX for v in group_map.values())

        if has_deduced:
            combined_cols = list(dict.fromkeys(master_df["Group_Sex"].dropna()))
        else:
            combined_cols = list(dict.fromkeys(
                f"{v.get('group', '')}_{v.get('sex', '')}" if isinstance(v, dict) else v
                for v in group_map.values()
            ))

        for data_key, target_folder, label in [
            ('Top_Val', folder_top, label_top),
            ('Bottom_Val', folder_bottom, label_bottom),
        ]:
            write_anat_pivot(master_df, 'Group_Sex', data_key, target_folder, label,
                             f"Mixed_{age_clean}", combined_cols)
            for sex_label in ['Male', 'Female']:
                sex_sub = os.path.join(target_folder, sex_label)
                os.makedirs(sex_sub, exist_ok=True)
                sex_df = master_df[master_df['Sex'] == sex_label].copy()
                if sex_df.empty:
                    continue
                if has_deduced:
                    sex_groups = list(dict.fromkeys(sex_df["Group"].dropna()))
                else:
                    sex_groups = list(dict.fromkeys(
                        v.get('group', '') for v in group_map.values()
                        if isinstance(v, dict) and v.get('sex') == sex_label
                    ))
                write_anat_pivot(sex_df, 'Group', data_key, sex_sub, label,
                                 f"{sex_label}_{age_clean}", sex_groups)
    else:
        expected_groups = list(group_map.values())
        for data_key, target_folder, label in [
            ('Top_Val', folder_top, label_top),
            ('Bottom_Val', folder_bottom, label_bottom),
        ]:
            write_anat_pivot(master_df, 'Group', data_key, target_folder, label,
                             f"{sex}_{age_clean}", expected_groups)


# ===========================================================
# PART 5: PIPELINE EXECUTION & DASHBOARD INTEGRATION
# ===========================================================

def generate_auto_master(output_folder, study_name, sex, age, bone):
    """
    Creates the Master Excel template automatically if it doesn't exist.
    Filename constraint: [StudyName]_[Sex]_[Age]_[Bone].xlsx
    """
    os.makedirs(output_folder, exist_ok=True)

    filename = f"{study_name}_{sex}_{age}_{bone}.xlsx"
    master_path = os.path.join(output_folder, filename)

    if not os.path.exists(master_path):
        print(f"Master file not found. Auto-generating template: {filename}")

        # Headers specifically aligned to sync_data_to_master search logic
        headers = [
            "Mouse Code",
            f"Avg. {bone} Length",
            f"Avg. {bone} Diameter",
            f"Avg. {bone} Thickness",
            "Maximum Load",
            "Stiffness",
            "Energy to Failure",
            "Displacement at Failure"
        ]

        df = pd.DataFrame(columns=headers)
        df.to_excel(master_path, index=False)
        print("Auto-generation complete.")
    else:
        print(f"Master file already exists, bypassing generation: {filename}")

    return master_path


def execute_single_study_pipeline(data_folder, master_path, measurement_path, csv_out_dir, group_map, study_name,
                                   bone='Bone', sex='Sex', age='Age',
                                   anatomical_diameters=False, measurement_sheet=None,
                                   bone_filter=None):
    root_path = Path(data_folder)

    analysis_excel = run_batch_bending_analysis(str(root_path), group_map)
    if not analysis_excel:
        print("Execution halted: No valid data files were evaluated.")
        return False

    if os.path.exists(master_path):
        sync_data_to_master(analysis_excel, master_path,
                            measurement_path, group_map, bone_filter=bone_filter)

        try:
            df_final = pd.read_excel(master_path)
            master_dir = os.path.dirname(master_path)

            compiled_filename = f"{study_name.replace(' ', '_')}_{bone}_Master_Compiled.csv"
            compiled_path = os.path.join(master_dir, compiled_filename)

            df_final.to_csv(compiled_path, index=False, sep=",")
            print(f"Main master file compiled and saved to: {compiled_path}")
        except Exception as e:
            print(f"Failed to generate compiled master copy: {e}")

        try:
            generate_segregated_csvs(master_path, csv_out_dir, group_map)
        except Exception as e:
            print(f"Error compiling side-by-side metrics: {e}")

    if anatomical_diameters:
        try:
            parse_anatomical_diameters_single_study(
                measurement_path, measurement_sheet, csv_out_dir, bone, sex, age, group_map,
                bone_filter=bone_filter)
        except Exception as e:
            print(f"Anatomical diameter parsing error: {e}")

    return True


def run_workflow(inputs):
    """
    Entry point for the Main Dashboard to execute the Single Study pipeline.
    """
    study_name = inputs.get('study_name')
    data_folder = inputs.get('data_folder')
    measurement_path = inputs.get('measurement_path')
    csv_out_dir = inputs.get('csv_out_dir')
    group_map = inputs.get('group_map', {})

    sex = inputs.get('sex', 'Sex')
    age = inputs.get('age', 'Age')
    bone = inputs.get('bone', 'Bone')
    anatomical_diameters = inputs.get('anatomical_diameters', False)
    measurement_sheet = inputs.get('measurement_sheet', None)
    auto_csv_subfolder = inputs.get('auto_csv_subfolder', True)

    print(f"Starting Single Study Workflow for: {study_name}")

    if measurement_path and os.path.exists(measurement_path):
        master_output_target = os.path.dirname(measurement_path)
    else:
        master_output_target = csv_out_dir

    # Detect bones from measurement file; fall back to configured bone if none found
    detected_bones = _detect_bones_from_measurement(measurement_path, measurement_sheet)
    bones_to_run = detected_bones if detected_bones else [bone]
    use_bone_subdirs = bool(detected_bones)

    all_success = True
    for bone_type in bones_to_run:
        print(f"\n--- Running pipeline for bone: {bone_type} ---")

        if use_bone_subdirs:
            bone_csv_dir = os.path.join(csv_out_dir, f"{study_name}_{bone_type}_CSVFiles")
        elif auto_csv_subfolder:
            bone_csv_dir = os.path.join(csv_out_dir, f"{study_name}_CSVFiles")
        else:
            bone_csv_dir = csv_out_dir
        bone_filter = bone_type if use_bone_subdirs else None

        master_path = generate_auto_master(
            master_output_target, study_name, sex, age, bone_type)

        success = execute_single_study_pipeline(
            data_folder=data_folder,
            master_path=master_path,
            measurement_path=measurement_path,
            csv_out_dir=bone_csv_dir,
            group_map=group_map,
            study_name=study_name,
            bone=bone_type,
            sex=sex,
            age=age,
            anatomical_diameters=anatomical_diameters,
            measurement_sheet=measurement_sheet,
            bone_filter=bone_filter,
        )
        all_success = all_success and success

    return all_success
