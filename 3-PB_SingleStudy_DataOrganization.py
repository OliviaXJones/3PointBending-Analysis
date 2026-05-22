import glob
import json
import os
import re
import sys
from datetime import datetime
from io import StringIO
from pathlib import Path

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

from PyQt6.QtCore import pyqtSignal
from PyQt6.QtWidgets import (
    QMainWindow, QDialog, QWidget, QVBoxLayout, QHBoxLayout,
    QFormLayout, QGroupBox, QLabel, QLineEdit, QComboBox,
    QPushButton, QFileDialog, QMessageBox, QApplication, QInputDialog,
    QTableWidget, QTableWidgetItem, QHeaderView

)

matplotlib.use("Agg")

# ===========================================================
# STUDY CONFIGURATION & AUTOMATIC STATE MANAGER
# ===========================================================
CONFIG_JSON_FILENAME = "studies_config.json"

DEFAULT_FALLBACK_CONFIG = {
    "IFS+SHP099+Medigel 2026": {
        "raw_data_root": r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending\Force-Displacement Raw Files\IFS+SHP099+Medigel_LFemur_051226",
        "master_file": r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending\IFS+SHP099+Medigel_LFemurMaster.xlsx",
        "measurement_file": r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending\Measurement Files\IFS+SHP099+Medigel_LFemur_051226.xlsx",
        "output_folder": r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending\IFS+SHP99+Medigel 2026",
        "group_map": {"CV": "Control + Medigel", "PV": "IFS + Medigel", "PS": "IFS + SHP Medigel"},
    }
}

ACTIVE_STUDY_NAME = "IFS+SHP099+Medigel 2026"
ACTIVE_GROUP_MAP = DEFAULT_FALLBACK_CONFIG["IFS+SHP099+Medigel 2026"]["group_map"]


def load_full_dictionary():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, CONFIG_JSON_FILENAME)
    if os.path.exists(json_path):
        try:
            with open(json_path, "r") as f:
                data = json.load(f)
                if data and isinstance(data, dict):
                    if "raw_data_root" in data:
                        return {"IFS+SHP099+Medigel 2026": data}
                    return data
        except Exception as e:
            print(f"Notice reading dictionary JSON: {e}")
    return DEFAULT_FALLBACK_CONFIG


def save_single_study_update(study_name, raw_dir, master_path, meas_path, out_dir):
    full_data = load_full_dictionary()

    existing_map = full_data.get(study_name, {}).get(
        "group_map") or full_data.get(study_name, {}).get("cohort_information")
    if not existing_map:
        existing_map = DEFAULT_FALLBACK_CONFIG["IFS+SHP099+Medigel 2026"]["group_map"]

    full_data[study_name] = {
        "raw_data_root": raw_dir,
        "master_file": master_path,
        "measurement_file": meas_path,
        "output_folder": out_dir,
        "group_map": existing_map
    }

    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, CONFIG_JSON_FILENAME)
    try:
        with open(json_path, "w") as f:
            json.dump(full_data, f, indent=4)
    except Exception as e:
        print(f"Could not automatically save layout update: {e}")


FILE_GLOB_PATTERN = "*.txt"
SAVE_PNG_DPI = 100
TOE_LOAD_FRACTION = 0.05
LINEAR_WINDOW_POINTS = 90
MIN_R2 = 0.995


def parse_id_and_group(text_string, group_map):
    if not text_string or pd.isna(text_string):
        return None, None, None

    text_string = str(text_string).strip()
    prefixes = list(group_map.keys())

    if not prefixes:
        return None, None, None

    pattern = rf"\b({'|'.join(prefixes)})\s*[-_]?\s*(\d+)"
    match = re.search(pattern, text_string, re.IGNORECASE)

    if match:
        pref = match.group(1).upper()
        num_id = match.group(2)
        group_name = group_map.get(pref, pref)
        return pref, num_id, group_name

    return None, None, None


# ===========================================================
# PART 1: BENDING DATA ANALYZER FUNCTIONS
# ===========================================================


def read_bending_txt(filepath):
    with open(filepath, "r", errors="ignore") as f:
        lines = f.readlines()

    data_start = None
    data_end = None
    for i, line in enumerate(lines):
        if "<DATA>" in line:
            data_start = i + 1
        elif "<END DATA>" in line and data_start:
            data_end = i
            break
    if data_start is None or data_end is None:
        raise ValueError(f"No valid <DATA> section in {filepath}")

    data_lines = lines[data_start:data_end]
    header = None
    for i, line in enumerate(data_lines):
        parts = line.strip().split("\t")
        try:
            float(parts[0])
        except ValueError:
            header = parts
            data_lines = data_lines[i + 1:]
            break

    df = pd.read_csv(StringIO("".join(data_lines)),
                     sep="\t", names=header, engine="python")
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.dropna(subset=["Fz, N", "Position (z), mm"], how="any")
    if "Fx" in df.columns:
        df["Fz, N"] = df["Fx"]
    if "Position (z)" in df.columns:
        df["Position (z), mm"] = df["Position (z)"]
    df["Fz, N"] = -df["Fz, N"]
    return df


def dominant_linear_region(x, y, window=30, min_r2=0.995):
    best_len = 0
    best = (0, 0, 0, 0)
    n = len(x)

    if n < window:
        return 0, 0, 0, n

    for start in range(0, n - window, 2):
        end = start + window
        x_seg = x[start:end]
        y_seg = y[start:end]

        m, b = np.polyfit(x_seg, y_seg, 1)
        r = np.corrcoef(x_seg, y_seg)[0, 1]
        r2 = r**2

        if r2 >= min_r2:
            while end < n:
                next_end = min(end + 5, n)
                x_ext = x[start:next_end]
                y_ext = y[start:next_end]
                m_ext, b_ext = np.polyfit(x_ext, y_ext, 1)
                r_ext = np.corrcoef(x_ext, y_ext)[0, 1]
                if (r_ext**2) < min_r2:
                    break
                m, b, end = m_ext, b_ext, next_end

            if (end - start) > best_len:
                best_len = end - start
                best = (m, b, start, end)

    return best


def run_batch_bending_analysis(input_folder, group_map):
    print(f"\n--- Processing Raw Data in Folder: {input_folder} ---")
    date_str = datetime.now().strftime("%m%d%y")

    plot_folder = os.path.join(input_folder, "Fz_Displacement_Analysis")
    os.makedirs(plot_folder, exist_ok=True)

    excel_filename = f"Fz_Displacement_Analysis_{date_str}.xlsx"
    txt_files = glob.glob(os.path.join(input_folder, FILE_GLOB_PATTERN))
    if not txt_files:
        return None

    results = []

    for path in txt_files:
        base_name = os.path.basename(path).replace(".txt", "")

        pref, num_id, g_name = parse_id_and_group(base_name, group_map)
        if not pref:
            print(
                f"Skipping file {base_name}: Doesn't match active group configuration prefixes.")
            continue

        try:
            df = read_bending_txt(path)
            displacement = df["Position (z), mm"].values
            load = df["Fz, N"].values

            load_smooth = np.convolve(load, np.ones(3) / 3, mode="same")
            DISPLACEMENT_LIMIT = 1.75
            valid_range_mask = displacement <= DISPLACEMENT_LIMIT

            load_for_peak = load_smooth[valid_range_mask]
            constrained_max = np.max(load_for_peak) if len(
                load_for_peak) > 0 else np.max(load_smooth)

            candidates = np.where(
                (load_smooth >= 0.5 * constrained_max) & (valid_range_mask))[0]
            max_idx = candidates[np.argmax(load[candidates])] if len(
                candidates) > 0 else np.argmax(load_smooth)

            max_load = load[max_idx]
            post_max_load = load[max_idx:]
            NEAR_ZERO_THRESHOLD = 0.5
            drop_threshold_value = max_load * 0.80

            zero_drop_indices = np.where(
                post_max_load <= NEAR_ZERO_THRESHOLD)[0]
            if len(zero_drop_indices) > 0:
                fail_idx = max_idx + zero_drop_indices[0]
            else:
                has_dropped_indices = np.where(
                    post_max_load < drop_threshold_value)[0]
                if len(has_dropped_indices) > 0:
                    search_start = has_dropped_indices[0]
                    search_load = post_max_load[search_start:]
                    diff_post = np.diff(np.convolve(
                        search_load, np.ones(3) / 3, mode="same"))
                    local_min_indices = np.where(diff_post >= 0)[0]
                    fail_idx = (
                        max_idx
                        + search_start
                        + (local_min_indices[0] if len(local_min_indices) > 0 else len(search_load) - 1)
                    )
                else:
                    fail_idx = len(load) - 1

            pre_max_load = load[:max_idx]
            toe_mask = pre_max_load >= (TOE_LOAD_FRACTION * max_load)
            toe_indices = np.where(toe_mask)[0]

            start_idx = toe_indices[0] if len(toe_indices) > 0 else 0
            adj_disp = displacement - displacement[start_idx]

            disp_at_failure = adj_disp[fail_idx]
            load_at_failure = load[fail_idx]

            disp_slope_candidates = displacement[:max_idx][toe_mask] - \
                displacement[start_idx]
            load_slope_candidates = pre_max_load[toe_mask]

            if len(disp_slope_candidates) < LINEAR_WINDOW_POINTS:
                stiffness, intercept, idx0, idx1 = np.nan, np.nan, 0, 0
            else:
                stiffness, intercept, idx0, idx1 = dominant_linear_region(
                    disp_slope_candidates, load_slope_candidates, window=LINEAR_WINDOW_POINTS
                )

            if hasattr(np, "trapezoid"):
                energy = np.trapezoid(
                    load[start_idx: fail_idx + 1], adj_disp[start_idx: fail_idx + 1])
            else:
                energy = np.trapz(
                    load[start_idx: fail_idx + 1], adj_disp[start_idx: fail_idx + 1])

            plt.figure(figsize=(7, 5))
            plt.plot(adj_disp, load, color="black", label="Data")
            if (idx1 - idx0) > 5 and np.isfinite(stiffness):
                x_lin = disp_slope_candidates[idx0:idx1]
                plt.plot(x_lin, stiffness * x_lin + intercept,
                         color="red", label=f"Stiffness: {stiffness:.2f} N/mm")

            plt.scatter(adj_disp[max_idx], max_load,
                        color="green", label="Max Load")
            plt.scatter(disp_at_failure, load_at_failure,
                        color="purple", label="Failure")
            plt.axvline(x=0, color="blue", linestyle="--", label="Toe End")
            plt.fill_between(
                adj_disp[start_idx: fail_idx + 1], load[start_idx: fail_idx + 1], alpha=0.2, color="orange"
            )
            plt.title(base_name)
            plt.legend()
            plt.savefig(os.path.join(
                plot_folder, f"{base_name}.png"), dpi=SAVE_PNG_DPI)
            plt.close()

            results.append(
                {
                    "Standardized_ID": f"{pref}{num_id}",
                    "Max_Load_N": round(max_load, 4),
                    "Stiffness_N_per_mm": round(stiffness, 4),
                    "Energy_to_Failure_Nmm": round(energy, 4),
                    "Displacement_at_Failure_mm": round(disp_at_failure, 4),
                }
            )
        except Exception as e:
            print(f"Error processing mechanics for {base_name}: {e}")

    if not results:
        return None

    output_excel_path = os.path.join(input_folder, excel_filename)
    pd.DataFrame(results).to_excel(output_excel_path, index=False)
    return output_excel_path


# ===========================================================
# PART 2: MASTER MERGING & CONSOLIDATION
# ===========================================================


def sync_data_to_master(analysis_excel_path, master_file, measurement_file, group_map):
    if not analysis_excel_path or not os.path.exists(analysis_excel_path):
        return

    df_mach = pd.read_excel(analysis_excel_path)
    df_meas = pd.DataFrame()

    if os.path.exists(measurement_file):
        try:
            df_meas = pd.read_excel(measurement_file)
        except Exception as e:
            print(f"Could not read measurement file: {e}")

    wb = openpyxl.load_workbook(master_file, keep_links=True)
    ws = wb.active

    headers = [str(ws.cell(row=1, column=c).value).strip()
               for c in range(1, ws.max_column + 1)]

    def find_col(name_options):
        for opt in name_options:
            if opt in headers:
                return headers.index(opt) + 1
        return None

    col_len = find_col(["Length", "Avg. Length", "Avg_Length"]) or 2
    col_diam = find_col(["Diameter", "Avg. Diameter", "Avg_Diameter"]) or 3
    col_thick = find_col(["Thickness", "Avg. Thickness", "Avg_Thickness"]) or 4
    col_max = find_col(["Maximum Load", "Max_Load", "Max_Load_N"]) or 5
    col_stiff = find_col(["Stiffness", "Stiffness_N_per_mm"]) or 6
    col_energy = find_col(
        ["Energy to Failure", "Energy_to_Failure", "Energy_to_Failure_Nmm"]) or 7
    col_disp = find_col(["Displacement at Failure",
                        "Displacement_at_Failure", "Displacement_at_Failure_mm"]) or 8

    for row in range(2, ws.max_row + 1):
        cell_val1 = ws.cell(row=row, column=1).value
        cell_val2 = ws.cell(row=row, column=2).value

        pref, num_id, _ = parse_id_and_group(cell_val1, group_map)
        if not pref:
            pref, num_id, _ = parse_id_and_group(cell_val2, group_map)

        if not pref:
            continue

        lookup_id = f"{pref}{num_id}"

        if not df_meas.empty:
            for idx, m_row in df_meas.iterrows():
                m_pref, m_num, _ = parse_id_and_group(m_row.iloc[0], group_map)
                if m_pref and f"{m_pref}{m_num}" == lookup_id:
                    ws.cell(
                        row=row, column=col_len).value = m_row.iloc[1] if df_meas.shape[1] > 1 else None
                    ws.cell(
                        row=row, column=col_diam).value = m_row.iloc[8] if df_meas.shape[1] > 8 else None
                    ws.cell(
                        row=row, column=col_thick).value = m_row.iloc[12] if df_meas.shape[1] > 12 else None
                    break

        mach_row = df_mach[df_mach["Standardized_ID"] == lookup_id]
        if not mach_row.empty:
            ws.cell(
                row=row, column=col_max).value = mach_row.iloc[0]["Max_Load_N"]
            ws.cell(
                row=row, column=col_stiff).value = mach_row.iloc[0]["Stiffness_N_per_mm"]
            ws.cell(
                row=row, column=col_energy).value = mach_row.iloc[0]["Energy_to_Failure_Nmm"]
            ws.cell(
                row=row, column=col_disp).value = mach_row.iloc[0]["Displacement_at_Failure_mm"]

    wb.save(master_file)


# ===========================================================
# PART 3: SEGREGATED SUB-STUDY CSV GENERATION (RE-PIVOTING)
# ===========================================================


def generate_segregated_csvs(master_path, export_dir, group_map):
    if master_path.endswith(".csv"):
        df = pd.read_csv(master_path)
    else:
        df = pd.read_excel(master_path)

    headers = [str(c).strip() for c in df.columns]

    def find_col_name(name_options):
        for opt in name_options:
            if opt in headers:
                return opt
            for h in headers:
                if opt.lower() in h.lower():
                    return h
        return None

    metric_mapping = {
        "Avg_Length": find_col_name(["Length", "Avg_Length", "Avg. Length"]),
        "Avg_Diameter": find_col_name(["Diameter", "Avg_Diameter", "Avg. Diameter"]),
        "Avg_Thickness": find_col_name(["Thickness", "Avg_Thickness", "Avg. Thickness"]),
        "Max_Load": find_col_name(["Max_Load", "Maximum Load", "Max Load", "Max_Load_N"]),
        "Stiffness": find_col_name(["Stiffness", "Stiffness_N_per_mm", "Stiffness (N/mm)"]),
        "Energy_to_Failure": find_col_name(["Energy", "Energy_to_Failure", "Energy to Failure"]),
        "Displacement_at_Failure": find_col_name(["Displacement", "Displacement_at_Failure"]),
    }

    parsed_rows = []

    for idx, row in df.iterrows():
        matched_pref, matched_num, matched_gname = None, None, None

        for col_idx in range(min(3, len(df.columns))):
            p, n, g = parse_id_and_group(row.iloc[col_idx], group_map)
            if p:
                matched_pref, matched_num, matched_gname = p, n, g
                break

        if not matched_pref:
            continue

        row_data = {"ID_Num": int(matched_num), "Group": matched_gname}

        for metric_key, actual_col in metric_mapping.items():
            row_data[metric_key] = row[actual_col] if actual_col else np.nan

        parsed_rows.append(row_data)

    if not parsed_rows:
        print("Notice: No matching cohort IDs could be extracted from master rows.")
        return

    df_parsed = pd.DataFrame(parsed_rows)
    expected_columns = list(group_map.values())
    os.makedirs(export_dir, exist_ok=True)

    for metric_name in metric_mapping.keys():
        if df_parsed[metric_name].isna().all():
            continue

        pivot_df = df_parsed.pivot_table(
            index="ID_Num", columns="Group", values=metric_name, aggfunc="first")
        pivot_df = pivot_df.reindex(columns=expected_columns)
        pivot_df = pivot_df.sort_index()

        out_csv_path = os.path.join(export_dir, f"{metric_name}.csv")

        with open(out_csv_path, "w", encoding="utf-8") as f:
            f.write("sep=,\n")

        pivot_df.to_csv(out_csv_path, mode="a", sep=",", index=True)
        print(f"Generated side-by-side table: {out_csv_path}")


# ===========================================================
# PART 4: PIPELINE EXECUTION ENGINE
# ===========================================================


def execute_single_study_pipeline(data_folder, master_path, measurement_path, csv_out_dir, group_map):
    root_path = Path(data_folder)

    analysis_excel = run_batch_bending_analysis(str(root_path), group_map)
    if not analysis_excel:
        print("Execution halted: No valid data files were evaluated.")
        return False

    if os.path.exists(master_path):
        sync_data_to_master(analysis_excel, master_path,
                            measurement_path, group_map)

        try:
            df_final = pd.read_excel(master_path)
            master_dir = os.path.dirname(master_path)

            compiled_filename = f"{ACTIVE_STUDY_NAME.replace(' ', '_')}_Master_Compiled.csv"
            compiled_path = os.path.join(master_dir, compiled_filename)

            df_final.to_csv(compiled_path, index=False, sep=",")
            print(f"Main master file compiled and saved to: {compiled_path}")
        except Exception as e:
            print(f"Failed to generate compiled master copy: {e}")

        try:
            generate_segregated_csvs(master_path, csv_out_dir, group_map)
        except Exception as e:
            print(f"Error compiling side-by-side metrics: {e}")

    return True


# ===========================================================
# PART 5: PYQT6 INTERFACE & WIZARD DESIGN (DARK MODE STYLED)
# ===========================================================

# Clean, high-contrast dark theme with enhanced text sizing
DARK_STYLE = """
    QMainWindow, QDialog {
        background-color: #1e1e1e;
    }
    QGroupBox {
        background-color: #252526;
        border: 1px solid #3f3f46;
        border-radius: 6px;
        margin-top: 10px;
        padding-top: 14px;
        font-weight: bold;
        font-size: 15px;
        color: #e4e4e7;
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        subcontrol-position: top left;
        left: 12px;
        padding: 0 4px;
    }
    QLabel {
        color: #d4d4d8;
        font-size: 14px;
    }
    QLineEdit {
        background-color: #2d2d30;
        color: #ffffff;
        border: 1px solid #3f3f46;
        border-radius: 4px;
        padding: 6px 10px;
        font-size: 14px;
    }
    QLineEdit:focus {
        border: 1px solid #007acc;
    }
    QComboBox {
        background-color: #2d2d30;
        color: #ffffff;
        border: 1px solid #3f3f46;
        border-radius: 4px;
        padding: 6px 28px 6px 10px;
        min-height: 26px;
        font-size: 14px;
    }
    QComboBox:focus {
        border: 1px solid #007acc;
    }
    QComboBox::drop-down {
        subcontrol-origin: padding;
        subcontrol-position: top right;
        width: 24px;
        border-left: 1px solid #3f3f46;
    }
    QComboBox QAbstractItemView {
        background-color: #2d2d30;
        color: #ffffff;
        selection-background-color: #007acc;
        selection-color: #ffffff;
        border: 1px solid #3f3f46;
        font-size: 14px;
    }
    QTableWidget {
        background-color: #2d2d30;
        color: #ffffff;
        border: 1px solid #3f3f46;
        gridline-color: #3f3f46;
        border-radius: 4px;
        font-size: 14px;
    }
    QTableWidget QLineEdit {
        background-color: #1e1e1e;
        color: #ffffff;
        border: 1px solid #007acc;
        padding: 2px 5px;
        font-size: 14px;
    }
    QTableWidget::item {
        padding: 6px;
    }
    QTableWidget::item {
        padding: 5px;
    }
    QTableWidget::item:focus {
        background-color: #3e3e42;
        color: #ffffff;
    }
    QHeaderView::section {
        background-color: #333337;
        color: #d4d4d8;
        padding: 4px;
        border: 1px solid #3f3f46;
        font-size: 13px;
        font-weight: bold;
    }
    QPushButton {
        background-color: #3e3e42;
        color: #ffffff;
        border: 1px solid #555555;
        border-radius: 4px;
        padding: 6px 14px;
        font-size: 14px;
    }
    QPushButton:hover {
        background-color: #4e4e52;
        border: 1px solid #007acc;
    }
    QPushButton:pressed {
        background-color: #2d2d30;
    }
    QMessageBox {
        background-color: #1e1e1e;
    }
"""


class CohortMappingWidget(QWidget):
    """
    A single-menu interface that places the sample prefix identifier 
    and treatment group on the same row for intuitive mapping.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(10)

        # Unified Mapping Table
        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(
            ["Filename Prefix (e.g., PS)", "Treatment Group (e.g., Paroxetine)"])

        # --- NEW VISIBILITY ENHANCEMENTS ---
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setAlternatingRowColors(True)

        # Give every row a comfortable, readable height (default is usually a cramped ~20px)
        self.table.verticalHeader().setDefaultSectionSize(36)
        # Make the header sections taller too
        self.table.horizontalHeader().setMinimumHeight(32)
        # ------------------------------------

        layout.addWidget(self.table)

        # Action Buttons for Rows
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)

        self.add_row_btn = QPushButton("+ Add Mapping Row")
        self.add_row_btn.clicked.connect(self.add_mapping_row)
        btn_layout.addWidget(self.add_row_btn)

        self.remove_row_btn = QPushButton("- Remove Selected")
        self.remove_row_btn.clicked.connect(self.remove_selected_row)
        btn_layout.addWidget(self.remove_row_btn)

        btn_layout.addStretch(1)
        layout.addLayout(btn_layout)

    def add_mapping_row(self, prefix="", treatment=""):
        row_idx = self.table.rowCount()
        self.table.insertRow(row_idx)

        prefix_item = QTableWidgetItem(prefix)
        treatment_item = QTableWidgetItem(treatment)

        self.table.setItem(row_idx, 0, prefix_item)
        self.table.setItem(row_idx, 1, treatment_item)

    def remove_selected_row(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            self.table.removeRow(current_row)

    def set_mapping(self, group_map):
        self.table.setRowCount(0)
        if not group_map:
            self.add_mapping_row()
            return

        for prefix, treatment in group_map.items():
            self.add_mapping_row(str(prefix), str(treatment))

    def get_mapping(self):
        group_map = {}
        for row in range(self.table.rowCount()):
            prefix_item = self.table.item(row, 0)
            treatment_item = self.table.item(row, 1)

            if prefix_item and treatment_item:
                prefix = prefix_item.text().strip().upper()
                treatment = treatment_item.text().strip()

                if prefix and treatment:
                    group_map[prefix] = treatment
        return group_map


class StudyConfigWizard(QDialog):
    config_saved = pyqtSignal(str)

    def __init__(self, parent=None, study_name="", config_path="studies_config.json"):
        super().__init__(parent)
        self.config_path = config_path
        self.study_name = study_name.strip()

        self.setWindowTitle(f"Configure Profile: {self.study_name}")
        self.resize(650, 600)
        self.setStyleSheet(DARK_STYLE)
        self.build_ui()

        self.full_data = load_full_dictionary()
        if self.study_name in self.full_data:
            self.prepopulate_existing_values(self.full_data[self.study_name])
        else:
            # Seed an empty row for brand new studies
            self.cohort_mapper.set_mapping({})

    def build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(12)

        dir_group = QGroupBox(
            f"Project Directories & Files: {self.study_name}")
        form_layout = QFormLayout(dir_group)
        form_layout.setSpacing(10)

        self.raw_data_root_edit = QLineEdit()
        form_layout.addRow("Raw Data Root:", self.create_path_row(
            self.raw_data_root_edit, is_folder=True))

        self.master_file_edit = QLineEdit()
        form_layout.addRow("Master File (.xlsx):", self.create_path_row(
            self.master_file_edit, is_folder=False))

        self.measurement_file_edit = QLineEdit()
        form_layout.addRow("Measurement File:", self.create_path_row(
            self.measurement_file_edit, is_folder=False))

        self.output_folder_edit = QLineEdit()
        form_layout.addRow("Output Folder:", self.create_path_row(
            self.output_folder_edit, is_folder=True))

        main_layout.addWidget(dir_group)

        cohort_group = QGroupBox("Cohort Information Mapping")
        cohort_layout = QVBoxLayout(cohort_group)

        # Insert our single-view table mapping widget here
        self.cohort_mapper = CohortMappingWidget()
        cohort_layout.addWidget(self.cohort_mapper)

        main_layout.addWidget(cohort_group)

        save_btn = QPushButton("Save Configuration Updates")
        save_btn.setStyleSheet(
            "background-color: #2e7d32; color: white; font-weight: bold; font-size: 15px; padding: 10px;")
        save_btn.clicked.connect(self.save_config)
        main_layout.addWidget(save_btn)

    def prepopulate_existing_values(self, profile):
        self.raw_data_root_edit.setText(profile.get("raw_data_root", ""))
        self.master_file_edit.setText(profile.get("master_file", ""))
        self.measurement_file_edit.setText(profile.get("measurement_file", ""))
        self.output_folder_edit.setText(profile.get("output_folder", ""))

        g_map = profile.get("group_map") or profile.get(
            "cohort_information") or {}
        self.cohort_mapper.set_mapping(g_map)

    def create_path_row(self, line_edit, is_folder):
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)
        layout.addWidget(line_edit)

        browse_btn = QPushButton("Browse")

        def browse_cmd():
            if is_folder:
                path = QFileDialog.getExistingDirectory(
                    self, "Select Directory")
            else:
                path, _ = QFileDialog.getOpenFileName(
                    self, "Select File", "", "Excel Files (*.xlsx);;All Files (*)")
            if path:
                line_edit.setText(path)

        browse_btn.clicked.connect(browse_cmd)
        layout.addWidget(browse_btn)
        return container

    def save_config(self):
        if not all([self.raw_data_root_edit.text(), self.master_file_edit.text(), self.output_folder_edit.text()]):
            QMessageBox.critical(
                self, "Error", "Required folder paths and files are missing.")
            return

        group_map_data = self.cohort_mapper.get_mapping()
        if not group_map_data:
            QMessageBox.critical(
                self, "Error", "You must input at least one complete group prefix code parameter.")
            return

        self.full_data[self.study_name] = {
            "raw_data_root": self.raw_data_root_edit.text(),
            "master_file": self.master_file_edit.text(),
            "measurement_file": self.measurement_file_edit.text(),
            "output_folder": self.output_folder_edit.text(),
            "group_map": group_map_data,
        }

        try:
            with open(self.config_path, "w") as f:
                json.dump(self.full_data, f, indent=4)
            QMessageBox.information(
                self, "Success", f"'{self.study_name}' written successfully.")
            self.config_saved.emit(self.study_name)
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Save Error",
                                 f"Could not write configuration data:\n{e}")


class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("3-Point Bending: Single Study Pipeline")
        self.setFixedSize(720, 560)
        self.setStyleSheet(DARK_STYLE)

        self.studies_dict = load_full_dictionary()
        self.build_ui()
        self.refresh_dropdown_options()

    def build_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(25, 15, 25, 15)
        main_layout.setSpacing(12)

        selector_group = QGroupBox("Study Selection")
        selector_layout = QHBoxLayout(selector_group)
        selector_layout.setContentsMargins(15, 12, 15, 12)

        self.study_selector = QComboBox()
        self.study_selector.setMinimumWidth(320)
        self.study_selector.setSizeAdjustPolicy(
            QComboBox.SizeAdjustPolicy.AdjustToContents)
        self.study_selector.currentIndexChanged.connect(self.on_study_changed)

        selector_layout.addWidget(self.study_selector)
        selector_layout.addStretch(1)
        main_layout.addWidget(selector_group)

        fields_group = QGroupBox("File Path Selections")
        fields_form = QFormLayout(fields_group)
        fields_form.setSpacing(10)

        self.path_raw_edit = QLineEdit()
        fields_form.addRow("Raw Data Folder:", self.create_main_browse_row(
            self.path_raw_edit, is_folder=True))

        self.path_master_edit = QLineEdit()
        fields_form.addRow("Master Excel File:", self.create_main_browse_row(
            self.path_master_edit, is_folder=False))

        self.path_meas_edit = QLineEdit()
        fields_form.addRow("Measurements File:", self.create_main_browse_row(
            self.path_meas_edit, is_folder=False))

        self.path_out_edit = QLineEdit()
        fields_form.addRow("CSV Export Folder:", self.create_main_browse_row(
            self.path_out_edit, is_folder=True))

        main_layout.addWidget(fields_group)

        btn_layout = QVBoxLayout()
        btn_layout.setSpacing(10)

        management_row = QHBoxLayout()
        management_row.setSpacing(10)

        add_config_btn = QPushButton("Add New Study")
        add_config_btn.setStyleSheet(
            "background-color: #007acc; color: white; padding: 10px; font-size: 14px; font-weight: bold;")
        add_config_btn.clicked.connect(self.add_new_study_profile)
        management_row.addWidget(add_config_btn)

        edit_config_btn = QPushButton("Edit Active Study")
        edit_config_btn.setStyleSheet(
            "background-color: #d8a000; color: #1e1e1e; padding: 10px; font-size: 14px; font-weight: bold;")
        edit_config_btn.clicked.connect(self.edit_active_study_profile)
        management_row.addWidget(edit_config_btn)

        btn_layout.addLayout(management_row)

        execute_btn = QPushButton("Execute Workflow")
        execute_btn.setStyleSheet(
            "background-color: #2e7d32; color: white; font-weight: bold; font-size: 15px; padding: 12px;")
        execute_btn.clicked.connect(self.run_pipeline)
        btn_layout.addWidget(execute_btn)

        main_layout.addLayout(btn_layout)

    def create_main_browse_row(self, line_edit, is_folder):
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)
        layout.addWidget(line_edit)

        browse_btn = QPushButton("Browse...")

        def browse_cmd():
            if is_folder:
                path = QFileDialog.getExistingDirectory(
                    self, "Select Directory")
            else:
                path, _ = QFileDialog.getOpenFileName(
                    self, "Select Excel File", "", "Excel Files (*.xlsx)")
            if path:
                line_edit.setText(path)
        browse_btn.clicked.connect(browse_cmd)
        layout.addWidget(browse_btn)
        return container

    def on_study_changed(self):
        global ACTIVE_STUDY_NAME, ACTIVE_GROUP_MAP
        selected = self.study_selector.currentText()
        if not selected or selected not in self.studies_dict:
            return

        ACTIVE_STUDY_NAME = selected
        profile = self.studies_dict[selected]

        self.path_raw_edit.setText(profile.get("raw_data_root", ""))
        self.path_master_edit.setText(profile.get("master_file", ""))
        self.path_meas_edit.setText(profile.get("measurement_file", ""))
        self.path_out_edit.setText(profile.get("output_folder", ""))

        ACTIVE_GROUP_MAP = profile.get(
            "group_map") or profile.get("cohort_information") or {}

    def refresh_dropdown_options(self, new_study_to_select=None):
        self.studies_dict = load_full_dictionary()

        self.study_selector.blockSignals(True)
        self.study_selector.clear()
        self.study_selector.addItems(list(self.studies_dict.keys()))
        self.study_selector.blockSignals(False)

        if new_study_to_select:
            self.study_selector.setCurrentText(new_study_to_select)
        elif ACTIVE_STUDY_NAME in self.studies_dict:
            self.study_selector.setCurrentText(ACTIVE_STUDY_NAME)
        elif self.studies_dict:
            self.study_selector.setCurrentIndex(0)

        self.on_study_changed()

    def add_new_study_profile(self):
        study_name, ok = QInputDialog.getText(
            self, "Create Profile", "Enter the unique name for the new study:")
        if ok and study_name.strip():
            wizard = StudyConfigWizard(self, study_name=study_name.strip())
            wizard.config_saved.connect(self.refresh_dropdown_options)
            wizard.exec()

    def edit_active_study_profile(self):
        active_profile = self.study_selector.currentText()
        if not active_profile:
            QMessageBox.warning(self, "Selection Error",
                                "No active workspace profile selected to edit.")
            return

        wizard = StudyConfigWizard(self, study_name=active_profile)
        wizard.config_saved.connect(self.refresh_dropdown_options)
        wizard.exec()

    def run_pipeline(self):
        self.setWindowTitle(
            f"Processing '{ACTIVE_STUDY_NAME}'... Please wait.")
        QApplication.processEvents()

        current_raw = self.path_raw_edit.text()
        current_master = self.path_master_edit.text()
        current_meas = self.path_meas_edit.text()
        current_out = self.path_out_edit.text()

        save_single_study_update(
            ACTIVE_STUDY_NAME, current_raw, current_master, current_meas, current_out)

        try:
            success = execute_single_study_pipeline(
                data_folder=current_raw,
                master_path=current_master,
                measurement_path=current_meas,
                csv_out_dir=current_out,
                group_map=ACTIVE_GROUP_MAP,
            )
            if success:
                QMessageBox.information(
                    self, "Success", f"Pipeline executed successfully for study: {ACTIVE_STUDY_NAME}")
            else:
                QMessageBox.warning(
                    self, "Incomplete", "No analysis calculations could be finalized.")
        except Exception as e:
            QMessageBox.critical(self, "Execution Error",
                                 f"An unhandled error occurred:\n{e}")
        finally:
            self.setWindowTitle("3-Point Bending: Single Study Pipeline")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
