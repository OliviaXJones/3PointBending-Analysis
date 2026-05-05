from pathlib import Path
from datetime import datetime
import glob
import openpyxl
import pandas as pd
import numpy as np
import os
from pzfx import write_pzfx

base_dir = r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending"
raw_data_root = os.path.join(
    base_dir, "Force-Displacement Raw Files", "FKBP5Null_Tibia_11226")
master_file = r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending\FKBP5_3-PointBendingTibiaMaster.xlsx"
measurement_file = r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending\Measurement Files\FKBP5Null_Tibia+Femur_11226.xlsx"
output_file = r"C:\Users\olivi\OneDrive - Medical University of South Carolina\3-Point Bending\FKBP5_3-PointBendingTibiaCompiled.csv"


def get_latest_analysis(search_path):
    root = Path(search_path)
    all_files = list(root.rglob("Fz_Displacement_Analysis_*.xlsx"))

    # --- DIAGNOSTIC PRINTS ---
    print(f"\n--- Checking directory: {search_path} ---")
    print(f"Found {len(all_files)} matching files:")
    for f in all_files:
        print(f"  - {f.name} (In: {f.parent.name})")
    # -------------------------

    if not all_files:
        raise FileNotFoundError(
            f"No files found matching 'Fz_Displacement_Analysis_*.xlsx' in {search_path}.")

    def sort_by_date(path_obj):
        try:
            # Extracts '040326' from 'Fz_Displacement_Analysis_040326'
            date_part = path_obj.stem.split('_')[-1]
            return datetime.strptime(date_part, "%m%d%y")
        except Exception as e:
            print(f"Warning: Could not parse date from {path_obj.name}: {e}")
            return datetime.min

    latest_path = max(all_files, key=sort_by_date)
    print(f"\n>>> SUCCESSFULLY SELECTED: {latest_path.name}\n")

    return str(latest_path)


def sync_data_to_master(all_analysis_files, master_file, measurement_file):
    # 1. Load Measurements (Tibia Length, etc.)
    # We still use this for the physical dimensions
    df_meas_all = pd.read_excel(measurement_file, sheet_name=None)

    # 2. Open the Master File
    wb = openpyxl.load_workbook(master_file)

    # 3. Create a mapping of Folder Name -> Latest File Path
    # This ensures if a folder has multiple dates, we only use the newest one
    folder_to_file_map = {}

    def get_date(path_obj):
        try:
            return datetime.strptime(path_obj.stem.split('_')[-1], "%m%d%y")
        except:
            return datetime.min

    for file_path in all_analysis_files:
        folder_name = file_path.parent.name  # e.g., "Wildtype F1"
        if folder_name not in folder_to_file_map:
            folder_to_file_map[folder_name] = file_path
        else:
            # Update only if this file is newer than the one already stored
            if get_date(file_path) > get_date(folder_to_file_map[folder_name]):
                folder_to_file_map[folder_name] = file_path

    # 4. Iterate through Master Sheets and pull from the mapped folder
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ["summary", "notes", "calculations"]:
            continue

        # Check if we have a data file for this specific sheet
        if sheet_name not in folder_to_file_map:
            print(f"Skipping {sheet_name}: No matching folder found.")
            continue

        print(
            f"Syncing Sheet '{sheet_name}' using file: {folder_to_file_map[sheet_name].name}")

        ws = wb[sheet_name]
        df_mach = pd.read_excel(
            folder_to_file_map[sheet_name])  # Sheet1 by default

        # Clean machine Mouse Codes
        df_mach['Mouse Code'] = df_mach['Filename'].str.replace(
            '.txt', '', regex=False)

        # Get physical measurements for this sheet
        df_meas = df_meas_all.get(sheet_name, pd.DataFrame())

        for row in range(2, ws.max_row + 1):
            for start_col in [1, 10]:  # A (Males) and J (Females)
                mouse_code = ws.cell(row=row, column=start_col).value
                if not mouse_code:
                    continue

                mouse_code = str(mouse_code).strip()

                # A. Write Physical Measurements (Length, Diam, Thick)
                if not df_meas.empty:
                    m_row = df_meas[df_meas.iloc[:, 0].astype(
                        str).str.strip() == mouse_code]
                    if not m_row.empty:
                        ws.cell(row=row, column=start_col +
                                1).value = m_row.iloc[0, 1]
                        ws.cell(row=row, column=start_col +
                                2).value = m_row.iloc[0, 8]
                        ws.cell(row=row, column=start_col +
                                3).value = m_row.iloc[0, 12]

                # B. Write Machine Data (Load, Stiffness, etc.)
                mach_row = df_mach[df_mach['Mouse Code'].astype(
                    str).str.strip() == mouse_code]
                if not mach_row.empty:
                    ws.cell(row=row, column=start_col +
                            4).value = mach_row.iloc[0]['Max_Load_N']
                    ws.cell(row=row, column=start_col +
                            5).value = mach_row.iloc[0]['Stiffness_N_per_mm']
                    ws.cell(row=row, column=start_col +
                            6).value = mach_row.iloc[0]['Energy_to_Failure_Nmm']
                    ws.cell(row=row, column=start_col +
                            7).value = mach_row.iloc[0]['Displacement_at_Failure_mm']

    wb.save(master_file)
    print("Master file update complete.")


def parse_mouse_code(code):
    ### Breaks down mouse code into Genotype, Age, Sex, and ID. ###
    try:
        parts = str(code).split('.')
        genotype = parts[0]
        age = parts[1]
        sex_id = parts[2]
        sex = "Male" if sex_id.startswith('M') else "Female"
        mouse_num = sex_id[1:]
        return genotype, age, sex, mouse_num
    except:
        return None, None, None, None


def generate_grouped_tables(df):
    metrics = [
        'Avg. Tibia Length', 'Avg. Tibia Diameter', 'Avg. Tibia Thickness',
        'Maximum Load', 'Stiffness', 'Energy to Failure', 'Displacement at Failure'
    ]
    unique_ages = df['Age_Extracted'].unique()

    # Custom sort order for sheet names/progeny groups
    # This ensures Wildtype sheets appear before Mutant, then Heterozygous
    sort_priority = ['Wildtype', 'Mutant', 'Heterozygous']

    # 1. Define and create sub-folders
    folder_genotype = os.path.join(base_dir, "Tibia_Analysis_By_Genotype")
    folder_lineage = os.path.join(base_dir, "Tibia_Analysis_By_Lineage")
    os.makedirs(folder_genotype, exist_ok=True)
    os.makedirs(folder_lineage, exist_ok=True)

    for metric in metrics:
        for age in unique_ages:
            for sex in ['Male', 'Female']:
                # Filter for Sex and Age first
                base_subset = df[(df['Sex_Extracted'] == sex) & (
                    df['Age_Extracted'] == age)].copy()
                if base_subset.empty:
                    continue

                clean_metric = metric.replace("_", " ").replace(".", "")

                # --- SET 1: Analysis By Genotype (ONLY 'Wildtype', 'Mutant', 'Heterozygous' sheets) ---
                # We filter the subset to ONLY include those three exact sheet names
                gen_subset = base_subset[base_subset['Progeny_Group'].isin(
                    sort_priority)].copy()

                if not gen_subset.empty:
                    table_genotype = gen_subset.pivot_table(
                        index='ID_Num', columns='Progeny_Group', values=metric)

                    # Ensure columns are in the specific WT -> Mutant -> Het order
                    existing_gen_cols = [
                        c for c in sort_priority if c in table_genotype.columns]
                    table_genotype = table_genotype.reindex(
                        columns=sort_priority)

                    name_genotype = f"{sex} {age}Wks {clean_metric}.csv"
                    table_genotype.to_csv(os.path.join(
                        folder_genotype, name_genotype))

                # --- SET 2: Analysis By Lineage (All data sheets, using original names) ---
                table_lineage = base_subset.pivot_table(
                    index='ID_Num', columns='Progeny_Group', values=metric)

                # Sort columns so Wildtype F1 comes before Mutant F1, etc.
                def lineage_sort(col_name):
                    for i, gen in enumerate(sort_priority):
                        if col_name.startswith(gen):
                            return (i, col_name)
                    return (99, col_name)

                sorted_lineage_cols = sorted(
                    table_lineage.columns, key=lineage_sort)
                table_lineage = table_lineage[sorted_lineage_cols]

                name_lineage = f"{sex} {age}Wks {clean_metric}.csv"
                table_lineage.to_csv(os.path.join(
                    folder_lineage, name_lineage))


def process_all_sheets(path):
    excel_data = pd.ExcelFile(path)
    sheet_names = excel_data.sheet_names
    all_compiled_data = []

    for sheet in sheet_names:
        # Ignore non-data sheets
        if sheet.lower() in ["summary", "notes", "calculations"]:
            continue

        print(f"Processing Sheet: {sheet}")
        raw_df = pd.read_excel(path, sheet_name=sheet, header=None, skiprows=1)

        headers = ['Mouse Code', 'Avg. Tibia Length', 'Avg. Tibia Diameter', 'Avg. Tibia Thickness',
                   'Maximum Load', 'Stiffness', 'Energy to Failure', 'Displacement at Failure']

        # Extract Male (A-H) and Female (J-Q)
        males = raw_df.iloc[:, 0:8].copy()
        females = raw_df.iloc[:, 9:17].copy()
        males.columns = headers
        females.columns = headers

        sheet_combined = pd.concat([males, females], ignore_index=True)
        sheet_combined = sheet_combined.dropna(subset=['Mouse Code'])

        # Tag with the actual Sheet Name (e.g. "Wildtype F1")
        sheet_combined['Progeny_Group'] = sheet
        all_compiled_data.append(sheet_combined)

    final_master_df = pd.concat(all_compiled_data, ignore_index=True)

    # Parse the codes
    final_master_df[['Genotype', 'Age_Extracted', 'Sex_Extracted', 'ID_Num']] = \
        final_master_df['Mouse Code'].apply(
            lambda x: pd.Series(parse_mouse_code(x)))

    # Run the generator
    generate_grouped_tables(final_master_df)

    # Final Master CSV
    final_master_df.to_csv(output_file, index=False)


if __name__ == "__main__":
    try:
        # Get ALL files instead of just one global "latest"
        root = Path(raw_data_root)
        all_analysis_files = list(root.rglob(
            "Fz_Displacement_Analysis_*.xlsx"))

        sync_data_to_master(all_analysis_files, master_file, measurement_file)
        process_all_sheets(master_file)

        print("\nWorkflow Complete!")
    except Exception as e:
        print(f"\nAn error occurred: {e}")
