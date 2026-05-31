import os
import sys
import glob
from io import StringIO
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
import matplotlib
import matplotlib.pyplot as plt

matplotlib.use("Agg")

# ===========================================================
# DEFAULT CONFIGURATION & CONSTANTS
# ===========================================================
DEFAULT_RAW_DATA_ROOT = r"F:\3-Point Bending\FKBP5 Genotyping 2026\FKBP5Null_Tibia_11226"
DEFAULT_TIBIA_MASTER_FILE = r"F:\3-Point Bending\FKBP5 Genotyping 2026\FKBP5_3-PointBendingTibiaMaster.xlsx"
DEFAULT_FEMUR_MASTER_FILE = r"F:\3-Point Bending\FKBP5 Genotyping 2026\FKBP5_3-PointBendingFemurMaster.xlsx"
DEFAULT_MEASUREMENT_FILE = r"F:\3-Point Bending\FKBP5 Genotyping 2026\FKBP5Null_Tibia+Femur_11226.xlsx"
DEFAULT_CSV_OUTPUT_DIR = r"F:\3-Point Bending\FKBP5 Genotyping 2026\FKBP5_CSVFiles"

FILE_GLOB_PATTERN = "*.txt"
SAVE_PNG_DPI = 100
TOE_LOAD_FRACTION = 0.05
LINEAR_WINDOW_POINTS = 90
MIN_R2 = 0.995


# ===========================================================
# PART 1: BENDING DATA ANALYZER FUNCTIONS
# ===========================================================

def read_bending_txt(filepath):
    with open(filepath, "r", errors="ignore") as f:
        lines = f.readlines()

    data_start = None
    data_end = None
    for i, line in enumerate(lines):
        if "<DATA>" in line:
            data_start = i + 1
        elif "<END DATA>" in line and data_start:
            data_end = i
            break
    if data_start is None or data_end is None:
        raise ValueError(f"No valid <DATA> section in {filepath}")

    data_lines = lines[data_start:data_end]
    header = None
    for i, line in enumerate(data_lines):
        parts = line.strip().split("\t")
        try:
            float(parts[0])
        except ValueError:
            header = parts
            data_lines = data_lines[i + 1:]
            break

    df = pd.read_csv(
        StringIO("".join(data_lines)), sep="\t", names=header, engine="python"
    )
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.dropna(subset=["Fz, N", "Position (z), mm"], how="any")
    if "Fx" in df.columns:
        df["Fz, N"] = df["Fx"]
    if "Position (z)" in df.columns:
        df["Position (z), mm"] = df["Position (z)"]
    df["Fz, N"] = -df["Fz, N"]
    return df


def dominant_linear_region(x, y, window=30, min_r2=0.995):
    best_len = 0
    best = (0, 0, 0, 0)
    n = len(x)

    if n < window:
        return 0, 0, 0, n

    for start in range(0, n - window, 2):
        end = start + window
        x_seg = x[start:end]
        y_seg = y[start:end]

        m, b = np.polyfit(x_seg, y_seg, 1)
        r = np.corrcoef(x_seg, y_seg)[0, 1]
        r2 = r**2

        if r2 >= min_r2:
            while end < n:
                next_end = min(end + 5, n)
                x_ext = x[start:next_end]
                y_ext = y[start:next_end]
                m_ext, b_ext = np.polyfit(x_ext, y_ext, 1)
                r_ext = np.corrcoef(x_ext, y_ext)[0, 1]
                if (r_ext**2) < min_r2:
                    break
                m, b, end = m_ext, b_ext, next_end

            if (end - start) > best_len:
                best_len = end - start
                best = (m, b, start, end)

    return best


def determine_bone_type(folder_path, fallback_bone):
    txt_files = glob.glob(os.path.join(folder_path, FILE_GLOB_PATTERN))
    femur_count = 0
    tibia_count = 0

    for f in txt_files:
        name_lower = os.path.basename(f).lower()
        if "femur" in name_lower:
            femur_count += 1
        elif "tibia" in name_lower:
            tibia_count += 1

    if femur_count > 0 and femur_count >= tibia_count:
        return "Femur"
    if tibia_count > 0 and tibia_count > femur_count:
        return "Tibia"

    full_path_lower = str(folder_path).lower()
    if "femur" in full_path_lower:
        return "Femur"
    if "tibia" in full_path_lower:
        return "Tibia"

    return fallback_bone


def run_batch_bending_analysis(input_folder):
    print(f"\nProcessing Raw Data in Folder: {input_folder}")
    date_str = datetime.now().strftime("%m%d%y")

    plot_folder = os.path.join(input_folder, "Fz_Displacement_Analysis")
    os.makedirs(plot_folder, exist_ok=True)

    excel_filename = f"Fz_Displacement_Analysis_{date_str}.xlsx"

    txt_files = glob.glob(os.path.join(input_folder, FILE_GLOB_PATTERN))
    if not txt_files:
        return None

    results = []

    for path in txt_files:
        try:
            df = read_bending_txt(path)
            displacement = df["Position (z), mm"].values
            load = df["Fz, N"].values

            load_smooth = np.convolve(load, np.ones(3) / 3, mode="same")

            DISPLACEMENT_LIMIT = 1.75
            valid_range_mask = displacement <= DISPLACEMENT_LIMIT

            load_for_peak = load_smooth[valid_range_mask]
            constrained_max = (
                np.max(load_for_peak)
                if len(load_for_peak) > 0
                else np.max(load_smooth)
            )

            candidates = np.where(
                (load_smooth >= 0.5 * constrained_max) & (valid_range_mask)
            )[0]
            max_idx = (
                candidates[np.argmax(load[candidates])]
                if len(candidates) > 0
                else np.argmax(load_smooth)
            )

            max_load = load[max_idx]
            disp_at_max = displacement[max_idx]

            post_max_load = load[max_idx:]
            NEAR_ZERO_THRESHOLD = 0.5
            drop_threshold_value = max_load * 0.80

            zero_drop_indices = np.where(
                post_max_load <= NEAR_ZERO_THRESHOLD)[0]
            if len(zero_drop_indices) > 0:
                fail_idx = max_idx + zero_drop_indices[0]
            else:
                has_dropped_indices = np.where(
                    post_max_load < drop_threshold_value)[0]
                if len(has_dropped_indices) > 0:
                    search_start = has_dropped_indices[0]
                    search_load = post_max_load[search_start:]
                    diff_post = np.diff(
                        np.convolve(search_load, np.ones(3) / 3, mode="same")
                    )
                    local_min_indices = np.where(diff_post >= 0)[0]
                    fail_idx = (
                        max_idx
                        + search_start
                        + (
                            local_min_indices[0]
                            if len(local_min_indices) > 0
                            else len(search_load) - 1
                        )
                    )
                else:
                    fail_idx = len(load) - 1

            pre_max_disp = displacement[:max_idx]
            pre_max_load = load[:max_idx]

            toe_mask = pre_max_load >= (TOE_LOAD_FRACTION * max_load)
            toe_indices = np.where(toe_mask)[0]

            start_idx = toe_indices[0] if len(toe_indices) > 0 else 0
            adj_disp = displacement - displacement[start_idx]

            adj_max_disp = adj_disp[max_idx]
            adj_fail_disp = adj_disp[fail_idx]

            disp_at_failure = adj_fail_disp
            load_at_failure = load[fail_idx]

            disp_slope_candidates = (
                pre_max_disp[toe_mask] - displacement[start_idx])
            load_slope_candidates = pre_max_load[toe_mask]

            if len(disp_slope_candidates) < LINEAR_WINDOW_POINTS:
                stiffness = np.nan
                intercept = np.nan
                idx0, idx1 = 0, 0
            else:
                stiffness, intercept, idx0, idx1 = dominant_linear_region(
                    disp_slope_candidates,
                    load_slope_candidates,
                    window=LINEAR_WINDOW_POINTS,
                )

            if hasattr(np, 'trapezoid'):
                energy = np.trapezoid(
                    load[start_idx: fail_idx + 1], adj_disp[start_idx: fail_idx + 1])
            else:
                energy = np.trapz(
                    load[start_idx: fail_idx + 1], adj_disp[start_idx: fail_idx + 1])

            plt.figure(figsize=(7, 5))
            plt.plot(adj_disp, load, color="black", label="Data")
            if (idx1 - idx0) > 5 and np.isfinite(stiffness):
                x_lin = disp_slope_candidates[idx0:idx1]
                plt.plot(
                    x_lin,
                    stiffness * x_lin + intercept,
                    color="red",
                    label=f"Stiffness: {stiffness:.2f} N/mm",
                )

            plt.scatter(adj_max_disp, max_load,
                        color="green", label="Max Load")
            plt.scatter(adj_fail_disp, load_at_failure,
                        color="purple", label="Failure")
            plt.axvline(x=0, color="blue", linestyle="--",
                        label="Toe End (Start)")
            plt.fill_between(
                adj_disp[start_idx: fail_idx + 1],
                load[start_idx: fail_idx + 1],
                alpha=0.2,
                color="orange",
            )
            plt.title(os.path.basename(path))
            plt.legend()
            plt.savefig(
                os.path.join(plot_folder, os.path.basename(
                    path).replace(".txt", ".png")),
                dpi=SAVE_PNG_DPI,
            )
            plt.close()

            results.append(
                {
                    "Filename": os.path.basename(path),
                    "Max_Load_N": round(max_load, 4),
                    "Stiffness_N_per_mm": round(stiffness, 4),
                    "Energy_to_Failure_Nmm": round(energy, 4),
                    "Displacement_at_Failure_mm": round(disp_at_failure, 4),
                }
            )

        except Exception as e:
            print(f"Error processing {os.path.basename(path)}: {e}")

    if not results:
        return None

    output_excel_path = os.path.join(input_folder, excel_filename)
    pd.DataFrame(results).to_excel(output_excel_path, index=False)
    print(f"Saved local summary sheet to: {output_excel_path}")
    return output_excel_path


# ===========================================================
# PART 2: MASTER MERGING & EXPORT FUNCTIONS
# ===========================================================

def sync_data_to_master(all_analysis_files, master_file, measurement_file, bone_target, fallback_bone, structure_type):
    df_meas_all = pd.read_excel(measurement_file, sheet_name=None)

    try:
        wb = openpyxl.load_workbook(master_file)
    except Exception as e:
        print(f"Warning: Could not open {master_file}. Check path.")
        return

    all_mach_dfs = []
    for f in all_analysis_files:
        try:
            all_mach_dfs.append(pd.read_excel(f))
        except Exception:
            pass

    df_mach_all = pd.concat(
        all_mach_dfs, ignore_index=True) if all_mach_dfs else pd.DataFrame()

    if not df_mach_all.empty:
        df_mach_all["Base_Code"] = df_mach_all["Filename"].str.replace(
            ".txt", "", regex=False).str.split('_').str[0].str.upper().str.strip()

    target_columns = [
        1, 10] if structure_type == "Split Genders (Male/Female)" else [1]

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ["summary", "notes", "calculations"]:
            continue

        ws = wb[sheet_name]
        df_meas = df_meas_all.get(sheet_name, pd.DataFrame())

        if not df_meas.empty:
            df_meas["Base_Code"] = df_meas.iloc[:, 0].astype(
                str).str.split('_').str[0].str.upper().str.strip()

        for row in range(2, ws.max_row + 1):
            for start_col in target_columns:
                mouse_code = ws.cell(row=row, column=start_col).value
                if not mouse_code:
                    continue

                base_master_code = str(mouse_code).strip().split('_')[
                    0].upper()

                if not df_meas.empty:
                    meas_mask = (df_meas["Base_Code"] == base_master_code)
                    potential_meas = df_meas[meas_mask]

                    m_row = pd.DataFrame()
                    if bone_target == "Femur":
                        m_row = potential_meas[potential_meas.iloc[:, 0].astype(
                            str).str.upper().str.contains("FEMUR")]
                    else:
                        tibia_explicit = potential_meas[potential_meas.iloc[:, 0].astype(
                            str).str.upper().str.contains("TIBIA")]
                        if not tibia_explicit.empty:
                            m_row = tibia_explicit
                        else:
                            m_row = potential_meas[~potential_meas.iloc[:, 0].astype(
                                str).str.upper().str.contains("FEMUR")]

                    if not m_row.empty:
                        ws.cell(row=row, column=start_col +
                                1).value = m_row.iloc[0, 1]
                        ws.cell(row=row, column=start_col +
                                2).value = m_row.iloc[0, 2]
                        ws.cell(row=row, column=start_col +
                                3).value = m_row.iloc[0, 3]

                if not df_mach_all.empty:
                    mach_mask = (df_mach_all["Base_Code"] == base_master_code)
                    potential_mach = df_mach_all[mach_mask]

                    mach_row = pd.DataFrame()
                    if bone_target == "Femur":
                        mach_row = potential_mach[potential_mach["Filename"].str.upper(
                        ).str.contains("FEMUR")]
                    else:
                        tibia_explicit = potential_mach[potential_mach["Filename"].str.upper(
                        ).str.contains("TIBIA")]
                        if not tibia_explicit.empty:
                            mach_row = tibia_explicit
                        else:
                            mach_row = potential_mach[~potential_mach["Filename"].str.upper(
                            ).str.contains("FEMUR")]

                    if not mach_row.empty:
                        ws.cell(row=row, column=start_col +
                                4).value = mach_row.iloc[0]["Max_Load_N"]
                        ws.cell(row=row, column=start_col +
                                5).value = mach_row.iloc[0]["Stiffness_N_per_mm"]
                        ws.cell(row=row, column=start_col +
                                6).value = mach_row.iloc[0]["Energy_to_Failure_Nmm"]
                        ws.cell(row=row, column=start_col +
                                7).value = mach_row.iloc[0]["Displacement_at_Failure_mm"]

    try:
        wb.save(master_file)
        print(f"{bone_target} Master file synchronization complete.")
    except PermissionError:
        raise PermissionError(
            f"Cannot save {master_file}. Please close it in Excel first.")


def parse_mouse_code(code):
    try:
        clean_code = str(code).split('_')[0].strip()
        parts = clean_code.split('.')
        genotype = parts[0]
        age = parts[1]
        sex_id = parts[2]
        sex = "Male" if sex_id.startswith('M') else "Female"
        mouse_num = sex_id[1:]
        return genotype, age, sex, mouse_num
    except:
        return None, None, None, None


def generate_grouped_tables(df, base_dir, bone_target):
    metrics = [
        f'Avg. {bone_target} Length', f'Avg. {bone_target} Diameter', f'Avg. {bone_target} Thickness',
        'Maximum Load', 'Stiffness', 'Energy to Failure', 'Displacement at Failure'
    ]

    fallback_metrics = [
        'Avg. Tibia Length', 'Avg. Tibia Diameter', 'Avg. Tibia Thickness',
        'Avg. Femur Length', 'Avg. Femur Diameter', 'Avg. Femur Thickness',
        'Maximum Load', 'Stiffness', 'Energy to Failure', 'Displacement at Failure'
    ]

    if 'Age_Extracted' not in df.columns:
        return

    unique_ages = df['Age_Extracted'].dropna().unique()
    sort_priority = ['Wildtype', 'Mutant', 'Heterozygous']

    folder_genotype = os.path.join(
        base_dir, f"{bone_target}_Analysis_By_Genotype")
    folder_lineage = os.path.join(
        base_dir, f"{bone_target}_Analysis_By_Lineage")
    os.makedirs(folder_genotype, exist_ok=True)
    os.makedirs(folder_lineage, exist_ok=True)

    active_metrics = [m for m in metrics if m in df.columns]
    if not active_metrics:
        active_metrics = [m for m in fallback_metrics if m in df.columns]

    for metric in active_metrics:
        df[metric] = pd.to_numeric(df[metric], errors='coerce')

    for metric in active_metrics:
        for age in unique_ages:
            for sex in ['Male', 'Female']:
                base_subset = df[(df['Sex_Extracted'] == sex) & (
                    df['Age_Extracted'] == age)].copy()
                if base_subset.empty:
                    continue

                clean_metric = metric.replace("_", " ").replace(".", "")

                gen_subset = base_subset[base_subset['Progeny_Group'].isin(
                    sort_priority)].copy()
                if not gen_subset.empty:
                    table_genotype = gen_subset.pivot_table(
                        index='ID_Num', columns='Progeny_Group', values=metric)
                    existing_gen_cols = [
                        c for c in sort_priority if c in table_genotype.columns]
                    table_genotype = table_genotype.reindex(
                        columns=existing_gen_cols)
                    name_genotype = f"{sex} {age}Wks {clean_metric}.csv"
                    table_genotype.to_csv(os.path.join(
                        folder_genotype, name_genotype))

                table_lineage = base_subset.pivot_table(
                    index='ID_Num', columns='Progeny_Group', values=metric)

                def lineage_sort(col_name):
                    for i, gen in enumerate(sort_priority):
                        if col_name.startswith(gen):
                            return (i, col_name)
                    return (99, col_name)

                sorted_lineage_cols = sorted(
                    table_lineage.columns, key=lineage_sort)
                table_lineage = table_lineage[sorted_lineage_cols]
                name_lineage = f"{sex} {age}Wks {clean_metric}.csv"
                table_lineage.to_csv(os.path.join(
                    folder_lineage, name_lineage))


def process_all_sheets(master_path, structure_type, csv_out_dir, bone_target):
    excel_data = pd.ExcelFile(master_path)
    all_compiled_data = []

    headers = ['Mouse Code', f'Avg. {bone_target} Length', f'Avg. {bone_target} Diameter', f'Avg. {bone_target} Thickness',
               'Maximum Load', 'Stiffness', 'Energy to Failure', 'Displacement at Failure']

    for sheet in excel_data.sheet_names:
        if sheet.lower() in ["summary", "notes", "calculations"]:
            continue

        if structure_type == "Split Genders (Male/Female)":
            raw_df = pd.read_excel(
                master_path, sheet_name=sheet, header=None, skiprows=1)

            if raw_df.shape[1] < 17:
                males = raw_df.iloc[:, 0:8].copy()
                males.columns = headers[:males.shape[1]]
                sheet_combined = males
            else:
                males = raw_df.iloc[:, 0:8].copy()
                females = raw_df.iloc[:, 9:17].copy()
                males.columns = headers
                females.columns = headers
                sheet_combined = pd.concat([males, females], ignore_index=True)
        else:
            sheet_combined = pd.read_excel(master_path, sheet_name=sheet)
            if 'Mouse Code' not in sheet_combined.columns:
                sheet_combined.rename(
                    columns={sheet_combined.columns[0]: 'Mouse Code'}, inplace=True)

        sheet_combined = sheet_combined.dropna(subset=['Mouse Code'])
        sheet_combined['Progeny_Group'] = sheet
        all_compiled_data.append(sheet_combined)

    if not all_compiled_data:
        return pd.DataFrame()

    final_master_df = pd.concat(all_compiled_data, ignore_index=True)
    final_master_df[['Genotype', 'Age_Extracted', 'Sex_Extracted', 'ID_Num']] = \
        final_master_df['Mouse Code'].apply(
            lambda x: pd.Series(parse_mouse_code(x)))

    generate_grouped_tables(final_master_df, csv_out_dir, bone_target)
    return final_master_df


# ===========================================================
# PART 3: ANATOMICAL DIAMETER PARSING
# ===========================================================

def parse_anatomical_diameters(measurement_file, base_output_dir):
    print(f"\nExtracting Structural Diameters from: {measurement_file}")
    if not os.path.exists(measurement_file):
        return

    try:
        xl = pd.ExcelFile(measurement_file)
        sort_priority = ['Wildtype', 'Mutant', 'Heterozygous']

        for bone_type in ["Femur", "Tibia"]:
            if bone_type == "Femur":
                label_top, label_bottom = "Anteroposterior_Femur", "Mediolateral_Femur"
            else:
                label_top, label_bottom = "Proximal_Tibia", "Distal_Tibia"

            all_data = []

            for sheet in xl.sheet_names:
                df = pd.read_excel(measurement_file, sheet_name=sheet)
                if df.empty:
                    continue

                for index, row in df.iterrows():
                    raw_code = str(row.iloc[0]).strip()
                    if not raw_code or raw_code.lower() == 'nan':
                        continue

                    if bone_type == "Femur" and "_Femur" not in raw_code:
                        continue
                    if bone_type == "Tibia" and "_Femur" in raw_code:
                        continue

                    clean_code = raw_code.split('_')[0]
                    parts = clean_code.split('.')
                    if len(parts) < 3:
                        continue

                    age_val, sex_id = parts[1], parts[2]
                    sex_full = 'Male' if sex_id.startswith('M') else 'Female'
                    id_num = sex_id[1:]

                    try:
                        group_ce = row.iloc[2:5].astype(float)
                        group_fh = row.iloc[5:8].astype(float)
                        avg_ce, avg_fh = group_ce.mean(), group_fh.mean()
                    except (ValueError, TypeError):
                        continue

                    if pd.isna(avg_ce) or pd.isna(avg_fh):
                        continue

                    all_data.append({
                        'ID_Num': id_num,
                        'Age_Extracted': age_val,
                        'Sex_Extracted': sex_full,
                        'Progeny_Group': sheet,
                        'Top_Val': max(avg_ce, avg_fh),
                        'Bottom_Val': min(avg_ce, avg_fh)
                    })

            if not all_data:
                continue

            master_df = pd.DataFrame(all_data)
            folder_top = os.path.join(base_output_dir, label_top)
            folder_bottom = os.path.join(base_output_dir, label_bottom)

            os.makedirs(folder_top, exist_ok=True)
            os.makedirs(folder_bottom, exist_ok=True)

            mapping = [('Top_Val', folder_top, label_top),
                       ('Bottom_Val', folder_bottom, label_bottom)]

            for data_key, target_folder, file_label in mapping:
                for age in master_df['Age_Extracted'].unique():
                    for sex in ['Male', 'Female']:
                        subset = master_df[(master_df['Sex_Extracted'] == sex) &
                                           (master_df['Age_Extracted'] == age)].copy()
                        if subset.empty:
                            continue

                        table = subset.pivot_table(
                            index='ID_Num', columns='Progeny_Group', values=data_key)

                        def lineage_sort(col_name):
                            for i, gen in enumerate(sort_priority):
                                if col_name.startswith(gen):
                                    return (i, col_name)
                                return (99, col_name)

                        sorted_cols = sorted(table.columns, key=lineage_sort)
                        table = table[sorted_cols]
                        file_name = f"{sex}_{age}Wks_{file_label}.csv"
                        table.to_csv(os.path.join(target_folder, file_name))

    except Exception as e:
        print(f"Error parsing diameters: {e}")


# ===========================================================
# PART 4: SYSTEM EXECUTION CONTROL PIPELINE
# ===========================================================

def execute_pipeline(data_folder, tibia_master, femur_master, measurement_path, csv_out_dir, structure_type, fallback_bone):
    root_path = Path(os.path.normpath(data_folder))
    parse_anatomical_diameters(measurement_path, csv_out_dir)

    subfolders_with_data = set()
    for txt_file in root_path.rglob(FILE_GLOB_PATTERN):
        subfolders_with_data.add(txt_file.parent)

    if not subfolders_with_data:
        raise FileNotFoundError(
            f"No text files matching '{FILE_GLOB_PATTERN}' found in subfolders of {data_folder}")

    bone_groups = {"Tibia": [], "Femur": []}

    for folder in subfolders_with_data:
        assigned_bone = determine_bone_type(folder, fallback_bone)
        bone_groups[assigned_bone].append(folder)
        run_batch_bending_analysis(str(folder))

    all_analysis_files = list(root_path.rglob(
        "Fz_Displacement_Analysis_*.xlsx"))

    if not all_analysis_files:
        raise FileNotFoundError(
            "Batch analysis failed to generate Excel summary files.")

    for bone_key, master_dest in [("Tibia", tibia_master), ("Femur", femur_master)]:
        if not bone_groups[bone_key]:
            continue

        sync_data_to_master(all_analysis_files, master_dest,
                            measurement_path, bone_key, fallback_bone, structure_type)

        try:
            compiled_clean = process_all_sheets(
                master_dest, structure_type, csv_out_dir, bone_key)
            if not compiled_clean.empty:
                # 1. Grab the directory where the master file lives
                master_dir = os.path.dirname(master_dest)

                # 2. Set the static filename without the date (this forces it to overwrite)
                csv_name = f"FKBP5_{bone_key}Master_Compiled.csv"

                # 3. Save it right next to the Master Excel file
                final_csv_path = os.path.join(master_dir, csv_name)
                compiled_clean.to_csv(final_csv_path, index=False)

        except Exception as e:
            print(f"Post-processing error on {bone_key}: {e}")

    return True


# ===========================================================
# PART 5: DASHBOARD INTEGRATION LAYER
# ===========================================================

def run_workflow(inputs):
    """
    Entry point for the Main Dashboard to execute the FKBP5 specific pipeline.
    Expects a dictionary with the necessary file paths and parameters.
    """
    # Extract variables from the dictionary provided by the Hub
    raw_folder = inputs.get('data_folder')
    tibia_path = inputs.get('tibia_master')
    femur_path = inputs.get('femur_master')
    meas_path = inputs.get('measurement_path')
    csv_dir = inputs.get('csv_out_dir')
    structure_type = inputs.get('structure_type')
    fallback_bone = inputs.get('fallback_bone')

    print(f"Starting FKBP5 Workflow for: {raw_folder}")

    # Call the existing execution pipeline
    success = execute_pipeline(
        data_folder=raw_folder,
        tibia_master=tibia_path,
        femur_master=femur_path,
        measurement_path=meas_path,
        csv_out_dir=csv_dir,
        structure_type=structure_type,
        fallback_bone=fallback_bone
    )

    return success
